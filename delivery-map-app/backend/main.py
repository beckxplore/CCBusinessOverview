from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from contextlib import asynccontextmanager
import clickhouse_connect
import os
from dotenv import load_dotenv
import logging
from typing import List, Dict, Any, Tuple, Union
import json
from pathlib import Path
import csv
import math
from typing import Optional
from datetime import date, datetime, timedelta
import httpx
from functools import lru_cache
from collections import defaultdict

from google.auth.exceptions import GoogleAuthError

from services.google_sheets import (
    GoogleSheetsClient,
    GoogleSheetsNotConfigured,
    is_configured as google_sheets_configured,
)
from services.sheet_data import fetch_sheet_metrics, fetch_raw_sheet_data
from services.sensitivity import compute_leader_sensitivity, compute_weekly_retention
from services.b2b_mcp_client import get_b2b_mcp_client, format_date_range
from services.b2b_purchase_price import get_b2b_purchase_price_service

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ClickHouse configuration
CLICKHOUSE_HOST = os.getenv('CLICKHOUSE_HOST')
CLICKHOUSE_PORT = int(os.getenv('CLICKHOUSE_PORT_STR', '8123'))
CLICKHOUSE_USER = os.getenv('CLICKHOUSE_USER')
CLICKHOUSE_PASSWORD = os.getenv('CLICKHOUSE_PASSWORD')
CLICKHOUSE_DATABASE = os.getenv('CLICKHOUSE_DATABASE')
CLICKHOUSE_SECURE = os.getenv('CLICKHOUSE_SECURE_STR', 'false').lower() == 'true'
CLICKHOUSE_VERIFY = os.getenv('CLICKHOUSE_VERIFY_STR', 'false').lower() == 'true'


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Lifespan events to warm up caches on startup."""
    try:
        logger.info("Starting up: warming up caches...")
        
        # 1. Warmup ClickHouse connection
        try:
            get_clickhouse_client()
        except Exception as e:
            logger.warning(f"ClickHouse warmup failed (non-fatal): {e}")

        # 2. Warmup CSV/File caches
        # These are LRU cached, so calling them once loads them into memory
        try:
            _load_product_alias_index()
            _load_per_product_volume_ratios()
            _load_leader_coordinate_map()
        except Exception as e:
             logger.warning(f"File cache warmup failed: {e}")

        # 3. Warmup Google Sheets if enabled
        if GOOGLE_SHEETS_ENABLED:
             try:
                 GoogleSheetsClient.get_instance()
             except Exception as e:
                 logger.warning(f"Failed to initialize Google Sheets client on startup: {e}")

        logger.info("Startup warmup complete")
    except Exception as e:
        logger.error(f"Startup process encountered an error: {e}")
    
    yield
    
    logger.info("Shutting down")


# Initialize FastAPI app
app = FastAPI(
    title="Delivery Map Analytics API",
    description="API for delivery data analytics and mapping",
    version="1.0.0",
    lifespan=lifespan
)

# CORS configuration - allow localhost for development and Vercel domains for production
ALLOWED_ORIGINS = os.getenv('ALLOWED_ORIGINS', '').split(',') if os.getenv('ALLOWED_ORIGINS') else []
# Default origins for local development
default_origins = [
    "http://localhost:5173",
    "http://127.0.0.1:5173",
    "http://localhost:5174",
    "http://127.0.0.1:5174",
    "http://localhost:5178",
    "http://127.0.0.1:5178",
    "http://localhost:3000"
]
# Combine default and environment origins, filter out empty strings
cors_origins = [origin.strip() for origin in default_origins + ALLOWED_ORIGINS if origin.strip()]

# Add CORS middleware - allow all origins for now (can be restricted later)
# Note: allow_credentials must be False when using allow_origins=["*"]
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow all origins for testing
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Frontend static files directory (for Docker deployment)
FRONTEND_DIST = Path(__file__).parent.parent / "frontend" / "dist"

def get_clickhouse_client():
    """Create a new ClickHouse client instance. Returns None if connection fails."""
    try:
        logger.info(f"Connecting to ClickHouse at {CLICKHOUSE_HOST}:{CLICKHOUSE_PORT}")
        client = clickhouse_connect.get_client(
            host=CLICKHOUSE_HOST,
            port=CLICKHOUSE_PORT,
            username=CLICKHOUSE_USER,
            password=CLICKHOUSE_PASSWORD,
            database=CLICKHOUSE_DATABASE,
            secure=CLICKHOUSE_SECURE,
            verify=CLICKHOUSE_VERIFY,
            connect_timeout=10,
            send_receive_timeout=60
        )
        # Test the connection
        client.query("SELECT 1")
        logger.info("Successfully connected to ClickHouse")
        return client
    except Exception as e:
        logger.warning(f"Failed to connect to ClickHouse: {str(e)}")
        return None

def parse_coordinates(coord_str: str) -> tuple[float, float]:
    """Parse POINT(lon lat) format coordinates."""
    try:
        if coord_str and coord_str.startswith('POINT('):
            coords = coord_str[6:-1]  # Remove 'POINT(' and ')'
            parts = coords.split(' ')
            if len(parts) == 2:
                return float(parts[1]), float(parts[0])  # [lat, lon]
    except Exception as e:
        logger.warning(f"Failed to parse coordinates: {coord_str}, error: {e}")
    return 0.0, 0.0

# ---------- Forecast helpers (CSV-based) ----------
PROJECT_ROOT = Path(__file__).resolve().parents[2]  # repo root
DATA_POINTS_DIR = PROJECT_ROOT / "data_points"

LOCAL_SHOP_CSV = DATA_POINTS_DIR / "Local shop price history.csv"
LOCAL_SHOP_XLSX = DATA_POINTS_DIR / "Local shop price history.xlsx"
SGL_ORDER_PRICE_CSV = DATA_POINTS_DIR / "SGL Order & Price History Data.csv"
COMMISSION_LOOKUP_CSV = DATA_POINTS_DIR / "COMMISSION_LOOKUP.csv"
PRODUCT_ALIASES_FILE = Path(__file__).resolve().parent / "data" / "product_aliases.json"
PER_PRODUCT_VOLUME_CSV = DATA_POINTS_DIR / "Weekly per product Volume normal vs SGL.csv"

GOOGLE_SHEETS_ENABLED = google_sheets_configured()

PROCUREMENT_SHEET_ID = os.getenv("GSHEET_PROCUREMENT_ID")
PROCUREMENT_SHEET_WORKSHEET = os.getenv("GSHEET_PROCUREMENT_WORKSHEET", "ProcurementCosts")

LOCAL_PRICE_SHEET_ID = os.getenv("GSHEET_LOCAL_PRICE_ID")
LOCAL_PRICE_SHEET_WORKSHEET = os.getenv("GSHEET_LOCAL_PRICE_WORKSHEET", "LocalShopPrices")

OPERATIONAL_COST_SHEET_ID = os.getenv("GSHEET_OPERATIONAL_COST_ID")
OPERATIONAL_COST_SHEET_WORKSHEET = os.getenv("GSHEET_OPERATIONAL_COST_WORKSHEET", "OperationalCosts")

DAILY_OPERATIONAL_COST_SHEET_ID = os.getenv("GSHEET_DAILY_OPERATIONAL_COST_ID")
DAILY_OPERATIONAL_COST_SHEET_WORKSHEET = os.getenv("GSHEET_DAILY_OPERATIONAL_COST_WORKSHEET", "Daily CP 1 & 2")

PRODUCT_METRICS_LOOKBACK_DAYS = int(os.getenv("PRODUCT_METRICS_LOOKBACK_DAYS", "30"))

BENCHMARK_API_URL = os.getenv("BENCHMARK_API_URL")
BENCHMARK_API_KEY = os.getenv("BENCHMARK_API_KEY")
BENCHMARK_LOCATION_GROUPS = [
    "farm",
    "distribution-center",
    "local-shops",
    "sunday-market",
    "supermarket",
    "ecommerce",
    "chipchip",
]

SELLING_PRICE_LOOKBACK_DAYS = int(os.getenv("SELLING_PRICE_LOOKBACK_DAYS", "30"))
FORECAST_WEEKLY_LOOKBACK_WEEKS = int(os.getenv("FORECAST_WEEKLY_LOOKBACK_WEEKS", "26"))

APRIL_DATA_START = date(2025, 4, 1)
WEEK_END_WEEKDAY = 3  # Thursday
WEEK_LENGTH_DAYS = 7


def _get_week_window(reference: Optional[date] = None) -> tuple[date, date]:
    """
    Get the analysis window.
    Defaults to a rolling 7-day window ending today to ensure data is current.
    """
    today = reference or datetime.utcnow().date()
    week_end = today
    week_start = week_end - timedelta(days=WEEK_LENGTH_DAYS - 1)
    
    if week_start < APRIL_DATA_START:
        week_start = APRIL_DATA_START
        
    return week_start, week_end


def _get_week_window_datetimes(reference: Optional[date] = None) -> tuple[date, date, datetime, datetime]:
    week_start, week_end = _get_week_window(reference)
    start_dt = datetime.combine(week_start, datetime.min.time())
    end_dt = datetime.combine(week_end + timedelta(days=1), datetime.min.time())
    return week_start, week_end, start_dt, end_dt




@lru_cache(maxsize=8)
def _get_sales_purchase_summary_for_window(start_key: str, end_key: str) -> dict[str, dict[str, Any]]:
    try:
        week_start = datetime.strptime(start_key, "%Y-%m-%d").date()
        week_end = datetime.strptime(end_key, "%Y-%m-%d").date()
    except ValueError:
        week_start, week_end = _get_week_window()

    df = fetch_raw_sheet_data()
    summary: dict[str, dict[str, Any]] = {}
    if df is None or df.empty:
        return summary

    for _, row in df.iterrows():
        raw_name = str(row.get("Product Name", "")).strip()
        if not raw_name:
            continue
        canonical = _normalize_product_name(raw_name)
        if not canonical:
            continue
        
        entry_date = row['date_dt'].date()
        purchase_price = float(row.get("PurchasingPrice", 0.0))
        selling_price = float(row.get("price", 0.0))
        quantity = float(row.get("final_volume_kg", 0.0))

        info = summary.setdefault(
            canonical,
            {
                "latest_purchase_price": None,
                "latest_purchase_date": None,
                "latest_selling_price": None,
                "latest_selling_date": None,
                "purchase_sum": 0.0,
                "purchase_count": 0,
                "selling_sum": 0.0,
                "selling_count": 0,
                "total_quantity": 0.0,
                "weekly_quantity": 0.0,
                "weekly_purchase_cost": 0.0,
                "weekly_sales_revenue": 0.0,
            },
        )

        if purchase_price > 0:
            info["purchase_sum"] += purchase_price
            info["purchase_count"] += 1
            latest_date = info["latest_purchase_date"]
            if latest_date is None or entry_date > latest_date:
                info["latest_purchase_price"] = purchase_price
                info["latest_purchase_date"] = entry_date

        if selling_price > 0:
            info["selling_sum"] += selling_price
            info["selling_count"] += 1
            latest_sell_date = info["latest_selling_date"]
            if latest_sell_date is None or entry_date > latest_sell_date:
                info["latest_selling_price"] = selling_price
                info["latest_selling_date"] = entry_date

        info["total_quantity"] += quantity

        if week_start <= entry_date <= week_end:
            info["weekly_quantity"] += quantity
            if purchase_price > 0:
                info["weekly_purchase_cost"] += purchase_price * quantity
            if selling_price > 0:
                info["weekly_sales_revenue"] += selling_price * quantity

    for info in summary.values():
        info["avg_purchase_price"] = (
            info["purchase_sum"] / info["purchase_count"] if info["purchase_count"] else None
        )
        info["avg_selling_price"] = (
            info["selling_sum"] / info["selling_count"] if info["selling_count"] else None
        )
        info.pop("purchase_sum", None)
        info.pop("purchase_count", None)
        info.pop("selling_sum", None)
        info.pop("selling_count", None)

    return summary


def _load_leader_coordinate_map() -> dict[str, Tuple[float, float]]:
    leaders_file = PROJECT_ROOT / "data_points" / "SGL_persona_leaders_unique.csv"
    coords: dict[str, Tuple[float, float]] = {}
    if not leaders_file.exists():
        return coords
    try:
        with leaders_file.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                phone = (row.get("phone") or "").strip()
                if not phone:
                    continue
                try:
                    lat = float(row.get("latest_delivery_latitude") or 0.0)
                    lon = float(row.get("latest_delivery_longitude") or 0.0)
                except (TypeError, ValueError):
                    continue
                if lat == 0.0 and lon == 0.0:
                    continue
                coords[phone] = (lat, lon)
    except Exception as exc:  # pylint: disable=broad-except
        logger.warning("Failed to load leader coordinate map: %s", exc)
    return coords


def _fetch_sheet_records(sheet_id: Optional[str], worksheet: str) -> Optional[list[dict[str, Any]]]:
    """Fetch Google Sheet records if configuration is available."""
    if not GOOGLE_SHEETS_ENABLED or not sheet_id:
        return None

    try:
        client = GoogleSheetsClient.get_instance()
        records = client.get_records(sheet_id, worksheet)
        return records
    except GoogleSheetsNotConfigured:
        logger.warning("Google Sheets not configured; falling back to local data for %s", worksheet)
        return None
    except GoogleAuthError as exc:
        logger.error("Google Sheets authentication failed: %s", exc)
        raise HTTPException(status_code=500, detail="Google Sheets authentication failed. Check service account credentials.")
    except Exception as exc:  # pylint: disable=broad-except
        logger.error("Failed to fetch data from Google Sheets (%s/%s): %s", sheet_id, worksheet, exc)
        return None


def _normalize_record_keys(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    for row in records:
        normalized_row = {str(k).strip().lower(): v for k, v in row.items() if k is not None}
        normalized.append(normalized_row)
    return normalized


def _parse_float(value: Any, default: float = 0.0) -> float:
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    try:
        text = str(value).strip().replace(",", "")
        if not text:
            return default
        return float(text)
    except Exception:  # pylint: disable=broad-except
        return default


def _sanitize_numeric_fields(record: dict[str, Any]) -> None:
    """Replace NaN or infinity float values with 0.0 to keep JSON serialization safe."""
    for key, value in record.items():
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            record[key] = 0.0


def _parse_datetime(value: Any) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    if not text:
        return None
    # Common formats across Google Sheets and CSV exports
    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%d/%m/%Y",
        "%Y/%m/%d",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None

def _load_products_from_xlsx(path: Path) -> Optional[list[str]]:
    try:
        import openpyxl  # lightweight dependency
    except Exception:
        return None
    try:
        wb = openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)
        ws = wb.active
        # Find header row and product column index
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(c).strip() if c is not None else '' for c in row]
            break
        if not headers:
            return []
        # Common header names
        candidates = {"product_name", "Product Name", "product", "name"}
        try:
            col_idx = next(i for i, h in enumerate(headers) if h in candidates)
        except StopIteration:
            # default to second column if classic format [date, product_name, price]
            col_idx = 1 if len(headers) >= 2 else 0
        products: set[str] = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None or len(row) == 0:
                continue
            cell = row[col_idx] if col_idx < len(row) else None
            name = str(cell).strip() if cell is not None else ''
            if name:
                products.add(name)
        return sorted(products)
    except FileNotFoundError:
        return None
    except Exception:
        return None


def _fetch_benchmark_price_map() -> dict[str, dict[str, Any]]:
    """Fetch benchmark prices from Supabase function and aggregate by product."""
    if not BENCHMARK_API_URL or not BENCHMARK_API_KEY:
        return {}

    today = datetime.utcnow().date()
    date_from = today - timedelta(days=6)  # inclusive 7-day window

    params = {
        "dateFrom": date_from.strftime("%Y-%m-%d"),
        "dateTo": today.strftime("%Y-%m-%d"),
        "frequency": "daily",
        "comparisonType": "avg",
        "locationGroups": ",".join(BENCHMARK_LOCATION_GROUPS),
    }

    headers = {
        "apikey": BENCHMARK_API_KEY,
        "Authorization": f"Bearer {BENCHMARK_API_KEY}",
        "Content-Type": "application/json",
    }

    try:
        with httpx.Client(timeout=15.0) as client:
            response = client.get(BENCHMARK_API_URL, params=params, headers=headers)
            response.raise_for_status()
            payload = response.json()
    except httpx.RequestError as exc:
        logger.warning("Benchmark API request failed: %s", exc)
        return {}
    except httpx.HTTPStatusError as exc:
        status = exc.response.status_code if exc.response is not None else "unknown"
        logger.warning("Benchmark API returned status %s: %s", status, exc)
        return {}
    except ValueError as exc:
        logger.warning("Benchmark API returned invalid JSON: %s", exc)
        return {}

    data_entries = payload.get("data") or []
    if not data_entries:
        return {}

    aggregates: dict[str, dict[str, Any]] = {}

    for entry in data_entries:
        product = (entry.get("product_name") or entry.get("product") or "").strip()
        price = _parse_float(entry.get("price"))
        if not product or price <= 0:
            continue

        product_stats = aggregates.setdefault(
            product,
            {
                "total_price": 0.0,
                "count": 0,
                "latest_date": None,
                "source": entry.get("source") or "supabase_benchmark",
            },
        )
        product_stats["total_price"] += price
        product_stats["count"] += 1

        last_checked = _parse_datetime(entry.get("date"))
        current_latest = product_stats["latest_date"]
        if last_checked and (current_latest is None or last_checked > current_latest):
            product_stats["latest_date"] = last_checked

    price_map: dict[str, dict[str, Any]] = {}
    for product, stats in aggregates.items():
        if stats["count"] == 0:
            continue
        avg_price = stats["total_price"] / stats["count"]
        latest_date = stats["latest_date"]
        price_map[product] = {
            "product_name": product,
            "price": round(avg_price, 2),
            "last_checked": latest_date.isoformat() if latest_date else None,
            "source": stats.get("source") or "supabase_benchmark",
        }

    return price_map


def _fetch_latest_selling_price_map() -> dict[str, float]:
    """
    Query ClickHouse for the latest selling price per product within a recent window.
    Falls back to an empty dict if ClickHouse is unavailable.
    """
    client = get_clickhouse_client()
    if not client:
        logger.warning("ClickHouse unavailable for latest selling prices")
        return {}

    query = f"""
    WITH toDateTime(now()) - INTERVAL {SELLING_PRICE_LOOKBACK_DAYS} DAY AS period_start
    SELECT
        any(coalesce(pn.name, pn_fallback.local_name, '')) AS product_name,
        argMax(gd.group_price, o.created_at) AS latest_selling_price
    FROM orders AS o
    JOIN groups_carts AS gc ON o.groups_carts_id = gc.id
    JOIN groups AS g ON gc.group_id = g.id
    JOIN group_deals AS gd ON g.group_deals_id = gd.id
    LEFT JOIN products AS prod ON gd.product_id = prod.id
    LEFT JOIN product_names AS pn ON prod.name_id = pn.id
    LEFT JOIN product_name_localizations AS pn_fallback ON pn.id = pn_fallback.product_name_id
        AND pn_fallback.language_code = 'en'
    WHERE o._peerdb_is_deleted = 0
      AND gc._peerdb_is_deleted = 0
      AND g._peerdb_is_deleted = 0
      AND gd._peerdb_is_deleted = 0
      AND o.status = 'COMPLETED' AND o.deleted_at IS NULL
      AND gc.status = 'COMPLETED' AND gc.deleted_at IS NULL
      AND g.status = 'COMPLETED' AND g.deleted_at IS NULL
      AND o.created_at >= period_start
    GROUP BY gd.product_id
    HAVING latest_selling_price > 0 AND product_name != ''
    """

    try:
        price_map: dict[str, float] = {}
        result = client.query(query)
        for row in result.named_results():
            product_name = row.get("product_name")
            price = row.get("latest_selling_price")
            if not product_name or price is None:
                continue
            normalized_name = str(product_name).strip().lower()
            if not normalized_name:
                continue
            price_map[normalized_name] = float(price)
        return price_map
    except Exception as exc:  # pylint: disable=broad-except
        logger.error("Failed to load latest selling prices from ClickHouse: %s", exc)
        return {}


@lru_cache(maxsize=1)
def _load_product_alias_index() -> dict[str, str]:
    alias_map: dict[str, str] = {}
    try:
        with PRODUCT_ALIASES_FILE.open("r", encoding="utf-8") as alias_file:
            records = json.load(alias_file)
        for record in records:
            canonical = str(record.get("canonical") or "").strip()
            if not canonical:
                continue
            canonical_lower = canonical.lower()
            alias_map[canonical_lower] = canonical
            for key in ("order_variants", "benchmark_variants", "distribution_variants"):
                for variant in record.get(key, []) or []:
                    if not variant:
                        continue
                    alias_map[str(variant).strip().lower()] = canonical
    except FileNotFoundError:
        logger.warning("Product aliases file not found at %s", PRODUCT_ALIASES_FILE)
    except Exception as exc:  # pylint: disable=broad-except
        logger.warning("Failed to load product aliases: %s", exc)
    return alias_map


def _normalize_product_name(name: str) -> str:
    cleaned = (name or "").strip()
    if not cleaned:
        return ""
    alias_map = _load_product_alias_index()
    return alias_map.get(cleaned.lower(), cleaned)


@lru_cache(maxsize=1)
def _load_per_product_volume_ratios() -> dict[str, float]:
    """
    Load average SGL volume ratios per product from the weekly per-product volume CSV.
    Returns a map of canonical product name -> sgl_ratio (0-1).
    """
    if not PER_PRODUCT_VOLUME_CSV.exists():
        return {}

    aggregates: dict[str, dict[str, float]] = defaultdict(lambda: {"sgl": 0.0, "total": 0.0})

    try:
        with PER_PRODUCT_VOLUME_CSV.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                product = (row.get("product_name") or "").strip()
                if not product:
                    continue
                canonical = _normalize_product_name(product)
                total = _parse_float(row.get("total_volume_kg"))
                sgl = _parse_float(row.get("sgl_volume_kg"))
                if total <= 0 or sgl < 0:
                    continue
                entry = aggregates[canonical]
                entry["total"] += total
                entry["sgl"] += min(sgl, total)
    except Exception as exc:  # pylint: disable=broad-except
        logger.warning("Failed to load per-product volume ratios from %s: %s", PER_PRODUCT_VOLUME_CSV, exc)
        return {}

    ratios: dict[str, float] = {}
    for product, values in aggregates.items():
        total = values["total"]
        sgl = values["sgl"]
        if total > 0:
            ratio = max(0.0, min(sgl / total, 1.0))
            ratios[product] = ratio

    return ratios


def _apply_volume_ratio_overrides(metrics: list[dict[str, Any]]) -> None:
    """
    Adjust SGL vs normal volume mix using the per-product ratio fallback when ClickHouse
    does not provide a usable breakdown (e.g., normal volume is missing or zero).
    """
    if not metrics:
        return

    ratio_map = _load_per_product_volume_ratios()
    if not ratio_map:
        return

    for entry in metrics:
        product_name = entry.get("product_name")
        if not product_name:
            continue
        total_volume = _parse_float(entry.get("total_volume_kg"))
        if total_volume <= 0:
            continue
        current_sgl = _parse_float(entry.get("sgl_volume_kg"))
        current_normal = _parse_float(entry.get("normal_volume_kg"))

        if current_normal > 0 and current_sgl < total_volume:
            # Already has a meaningful split; skip override.
            continue

        canonical = _normalize_product_name(str(product_name))
        ratio = ratio_map.get(canonical)
        if ratio is None:
            continue

        adjusted_sgl = round(total_volume * ratio, 4)
        adjusted_normal = max(total_volume - adjusted_sgl, 0.0)
        entry["sgl_volume_kg"] = adjusted_sgl
        entry["normal_volume_kg"] = adjusted_normal


def _fetch_latest_weekly_summary() -> Optional[dict[str, Any]]:
    """
    Retrieve the most recent completed week's order count and volume from ClickHouse.
    Returns None if ClickHouse is unavailable or no data is present.
    """
    client = get_clickhouse_client()
    if not client:
        logger.warning("ClickHouse unavailable for weekly summary")
        return None

    week_start, week_end, week_start_dt, week_end_dt = _get_week_window_datetimes()
    start_str = week_start_dt.strftime("%Y-%m-%d %H:%M:%S")
    end_str = week_end_dt.strftime("%Y-%m-%d %H:%M:%S")

    orders_query = f"""
    SELECT countDistinct(o.id) AS total_orders
    FROM orders AS o
    JOIN groups_carts AS gc ON o.groups_carts_id = gc.id
    JOIN groups AS g ON gc.group_id = g.id
    WHERE
        o._peerdb_is_deleted = 0
        AND gc._peerdb_is_deleted = 0
        AND g._peerdb_is_deleted = 0
        AND o.status = 'COMPLETED' AND o.deleted_at IS NULL
        AND gc.status = 'COMPLETED' AND gc.deleted_at IS NULL
        AND g.status = 'COMPLETED' AND g.deleted_at IS NULL
        AND o.created_at >= toDateTime('{start_str}')
        AND o.created_at <  toDateTime('{end_str}')
    """

    weekly_orders: Optional[int] = None
    try:
        orders_result = client.query(orders_query)
        for row in orders_result.named_results():
            weekly_orders = int(row.get("total_orders") or 0)
            break
    except Exception as exc:  # pylint: disable=broad-except
        logger.warning("Failed to fetch weekly orders summary: %s", exc)

    volume_query = f"""
    SELECT sum(gc.quantity) AS total_kg
    FROM groups_carts AS gc
    JOIN groups AS g ON gc.group_id = g.id
    JOIN orders AS o ON o.groups_carts_id = gc.id
    WHERE
        gc._peerdb_is_deleted = 0
        AND g._peerdb_is_deleted = 0
        AND o._peerdb_is_deleted = 0
        AND gc.status = 'COMPLETED' AND gc.deleted_at IS NULL
        AND g.status = 'COMPLETED' AND g.deleted_at IS NULL
        AND o.status = 'COMPLETED' AND o.deleted_at IS NULL
        AND o.created_at >= toDateTime('{start_str}')
        AND o.created_at <  toDateTime('{end_str}')
    """

    weekly_volume: Optional[float] = None
    try:
        volume_result = client.query(volume_query)
        for row in volume_result.named_results():
            total_kg_value = row.get("total_kg")
            if total_kg_value is not None:
                weekly_volume = float(total_kg_value)
            break
    except Exception as exc:  # pylint: disable=broad-except
        logger.warning("Failed to fetch weekly volume summary: %s", exc)

    if weekly_orders is None and weekly_volume is None:
        return None

    return {
        "week_start": week_start.isoformat(),
        "week_end": week_end.isoformat(),
        "orders": weekly_orders if weekly_orders is not None else None,
        "total_kg": weekly_volume if weekly_volume is not None else None,
        "source": "clickhouse",
    }


def load_local_shop_products() -> list[str]:
    """Distinct product names from local shop price benchmark file (xlsx preferred)."""
    sheet_records = _fetch_sheet_records(LOCAL_PRICE_SHEET_ID, LOCAL_PRICE_SHEET_WORKSHEET)
    if sheet_records:
        products: set[str] = set()
        for row in _normalize_record_keys(sheet_records):
            product = str(
                row.get("product_name")
                or row.get("product")
                or row.get("name")
                or ""
            ).strip()
            if product:
                products.add(product)
        if products:
            return sorted(products)

    # Prefer XLSX if available and readable
    if LOCAL_SHOP_XLSX.exists():
        products = _load_products_from_xlsx(LOCAL_SHOP_XLSX)
        if products is not None:
            return products
        # If openpyxl missing or parsing failed, fall back to CSV if present
    # Fallback to CSV
    if LOCAL_SHOP_CSV.exists():
        products: set[str] = set()
        try:
            with LOCAL_SHOP_CSV.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    name = (row.get("product_name") or row.get("Product Name") or "").strip()
                    if name:
                        products.add(name)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed reading local shop CSV: {e}")
        return sorted(products)
    raise HTTPException(status_code=500, detail="Local shop price benchmark file not found (.xlsx or .csv)")

def compute_product_shares(allowed_products: set[str]) -> dict:
    """Compute demand shares using aggregated product metrics."""
    metrics = load_product_metrics_data()
    totals: dict[str, float] = {}
    grand_total = 0.0

    for entry in metrics:
        product = entry.get("product_name")
        if not product or (allowed_products and product not in allowed_products):
            continue
        volume = float(entry.get("total_volume_kg", 0.0))
        if volume <= 0:
            continue
        totals[product] = totals.get(product, 0.0) + volume
        grand_total += volume

    if grand_total == 0:
        return {"shares": {}, "overall_sum": 0.0}

    shares = {product: volume / grand_total for product, volume in totals.items()}
    return {"shares": shares, "overall_sum": grand_total}


def load_local_shop_price_map() -> dict[str, dict[str, Any]]:
    """Return latest local shop price per product from Google Sheets or CSV."""
    benchmark_price_map = _fetch_benchmark_price_map()
    if benchmark_price_map:
        return benchmark_price_map

    price_map: dict[str, dict[str, Any]] = {}

    sheet_records = _fetch_sheet_records(LOCAL_PRICE_SHEET_ID, LOCAL_PRICE_SHEET_WORKSHEET)
    if sheet_records:
        for row in _normalize_record_keys(sheet_records):
            product = str(
                row.get("product_name")
                or row.get("product")
                or row.get("name")
                or ""
            ).strip()
            price = _parse_float(
                row.get("price")
                or row.get("local_shop_price")
                or row.get("shop_price")
            )
            if not product or price <= 0:
                continue
            last_checked = _parse_datetime(
                row.get("last_checked")
                or row.get("updated_at")
                or row.get("date")
            )
            prev = price_map.get(product)
            if not prev or (last_checked and last_checked > prev.get("last_checked_dt")):
                price_map[product] = {
                    "product_name": product,
                    "price": price,
                    "last_checked": last_checked.isoformat() if last_checked else None,
                    "last_checked_dt": last_checked,
                    "source": "google_sheets",
                }
        if price_map:
            # Remove helper datetime field before returning
            for entry in price_map.values():
                entry.pop("last_checked_dt", None)
            return price_map

    if LOCAL_SHOP_XLSX.exists():
        # TODO: optionally parse XLSX for richer data; for now continue to CSV fallback
        pass

    if LOCAL_SHOP_CSV.exists():
        try:
            with LOCAL_SHOP_CSV.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    product = (row.get("product_name") or "").strip()
                    if not product:
                        continue
                    price = _parse_float(row.get("Local shop price") or row.get("local_shop_price"))
                    if price <= 0:
                        continue
                    last_checked = _parse_datetime(row.get("date"))
                    prev = price_map.get(product)
                    if not prev or (last_checked and last_checked > _parse_datetime(prev.get("last_checked"))):
                        price_map[product] = {
                            "product_name": product,
                            "price": price,
                            "last_checked": last_checked.isoformat() if last_checked else row.get("date"),
                            "source": "csv",
                        }
        except Exception as exc:  # pylint: disable=broad-except
            logger.error("Failed to load local shop price CSV: %s", exc)

    return price_map


def load_product_costs_data() -> list[dict[str, Any]]:
    latest_prices = _fetch_latest_selling_price_map()
    
    # Use sheet window to ensure consistency
    sheet_result = fetch_sheet_metrics()
    if sheet_result and sheet_result.get("window"):
        week_start = sheet_result["window"]["start"]
        week_end = sheet_result["window"]["end"]
    else:
        week_start, week_end = _get_week_window()

    sales_summary = _get_sales_purchase_summary_for_window(week_start.isoformat(), week_end.isoformat())

    def _summary_for(product_name: str) -> Optional[dict[str, Any]]:
        canonical = _normalize_product_name(product_name)
        return sales_summary.get(canonical)

    sheet_records = _fetch_sheet_records(PROCUREMENT_SHEET_ID, PROCUREMENT_SHEET_WORKSHEET)
    source_rows: Optional[list[dict[str, Any]]] = None
    if sheet_records:
        source_rows = list(_normalize_record_keys(sheet_records))
    else:
        costs_file = DATA_POINTS_DIR / "PRODUCT_COSTS.csv"
        if not costs_file.exists():
            raise HTTPException(status_code=404, detail="Product costs file not found.")
        with costs_file.open("r", encoding="utf-8-sig", newline="") as f:
            source_rows = list(csv.DictReader(f))

    products: list[dict[str, Any]] = []
    for row in _normalize_record_keys(source_rows):
        product_name = str(
            row.get("product_name")
            or row.get("product")
            or row.get("name")
            or ""
        ).strip()
        if not product_name:
            continue
        canonical_name = _normalize_product_name(product_name)
        lookup_key = canonical_name.lower()
        latest_price = latest_prices.get(lookup_key)
        summary = _summary_for(product_name)

        purchase_override = None
        if summary:
            purchase_override = summary.get("latest_purchase_price") or summary.get("avg_purchase_price")
        selling_override = latest_price
        if selling_override is None and summary:
            selling_override = summary.get("latest_selling_price") or summary.get("avg_selling_price")

        products.append(
            {
                "product_name": canonical_name,
                "procurement_cost": _parse_float(
                    purchase_override
                    if purchase_override is not None
                    else row.get("procurement_cost")
                    or row.get("procurement")
                    or row.get("procurement_etb")
                ),
                "operational_cost": _parse_float(
                    row.get("operational_cost")
                    or row.get("operations_cost")
                    or row.get("operational_cost_per_kg")
                ),
                "sgl_commission": _parse_float(
                    row.get("sgl_commission")
                    or row.get("sgl_commission_etb")
                    or row.get("sgl_commission_per_kg")
                    or 0.0
                ),
                "regular_commission": _parse_float(
                    row.get("regular_commission")
                    or row.get("regular_commission_etb")
                    or row.get("regular_commission_per_kg")
                    or 0.0
                ),
                "selling_price": selling_override
                if selling_override is not None
                else _parse_float(row.get("selling_price") or row.get("price") or row.get("current_price")),
                "notes": str(row.get("notes") or row.get("comment") or "").strip(),
            }
        )

    existing_keys = {_normalize_product_name(item["product_name"]).lower() for item in products}
    for canonical, info in sales_summary.items():
        key = canonical.lower()
        if key in existing_keys:
            continue
        procurement_cost = info.get("latest_purchase_price") or info.get("avg_purchase_price") or 0.0
        selling_price = info.get("latest_selling_price") or info.get("avg_selling_price") or 0.0
        products.append(
            {
                "product_name": canonical,
                "procurement_cost": procurement_cost,
                "operational_cost": 0.0,
                "sgl_commission": 0.0,
                "regular_commission": 0.0,
                "selling_price": selling_price,
                "notes": "Auto-imported from Google Sheet history",
            }
        )

    return products


def load_operational_costs_data() -> list[dict[str, Any]]:
    sheet_records = _fetch_sheet_records(OPERATIONAL_COST_SHEET_ID, OPERATIONAL_COST_SHEET_WORKSHEET)
    if sheet_records:
        costs: list[dict[str, Any]] = []
        for row in _normalize_record_keys(sheet_records):
            category = str(row.get("cost_category") or row.get("category") or row.get("name") or "").strip()
            if not category:
                continue
            costs.append(
                {
                    "cost_category": category,
                    "cost_per_kg": _parse_float(
                        row.get("cost_per_kg")
                        or row.get("per_kg")
                        or row.get("value")
                        or row.get("cost")
                    ),
                    "description": str(row.get("description") or row.get("notes") or "").strip(),
                    "optimization_potential": str(
                        row.get("optimization_potential")
                        or row.get("potential")
                        or row.get("priority")
                        or "Low"
                    ).strip()
                    or "Low",
                }
            )
        if costs:
            return costs

    costs_file = DATA_POINTS_DIR / "OPERATIONAL_COSTS.csv"
    if not costs_file.exists():
        raise HTTPException(status_code=404, detail="Operational costs file not found.")

    costs = []
    try:
        with costs_file.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                costs.append(
                    {
                        "cost_category": row.get("cost_category", "").strip(),
                        "cost_per_kg": _parse_float(row.get("cost_per_kg")),
                        "description": row.get("description", "").strip(),
                        "optimization_potential": row.get("optimization_potential", "Low").strip(),
                    }
                )
    except Exception as exc:  # pylint: disable=broad-except
        raise HTTPException(status_code=500, detail=f"Failed to load operational costs: {exc}") from exc

    return costs


def load_daily_operational_costs_data() -> list[dict[str, Any]]:
    """Load daily operational costs from Google Sheet with date columns.
    
    Expected sheet structure:
    - Row 1: Headers (Day, Responsible, Data Source, Link, then date columns)
    - Rows with cost categories: "Warehouse costs per Kg", "Fulfilment costs per Kg", "Last Mile Costs per Kg"
    - Date columns contain daily cost values per kg
    
    Uses raw values instead of records to handle duplicate column headers.
    """
    from datetime import datetime
    
    if not GOOGLE_SHEETS_ENABLED or not DAILY_OPERATIONAL_COST_SHEET_ID:
        logger.warning("Daily operational costs sheet not configured")
        return []
    
    try:
        client = GoogleSheetsClient.get_instance()
        spreadsheet = client._client.open_by_key(DAILY_OPERATIONAL_COST_SHEET_ID)
        ws = spreadsheet.worksheet(DAILY_OPERATIONAL_COST_SHEET_WORKSHEET)
        
        # Get all values as raw data (handles duplicate headers)
        all_values = ws.get_all_values()
        if not all_values or len(all_values) < 2:
            logger.warning("Daily operational costs sheet is empty or has no data rows")
            return []
        
        # First row is headers
        headers = [str(h).strip() for h in all_values[0]]
        
        # Find date columns (skip first few columns: Day, Responsible, Data Source, Link)
        date_column_indices: list[tuple[int, date]] = []
        for idx, header in enumerate(headers):
            # Skip known non-date columns
            header_lower = header.lower()
            if header_lower in ["day", "responsible", "data source", "link", "data source (system,manual , asumption , average)"]:
                continue
            
            # Try to parse as date
            try:
                date_str = header.strip()
                if not date_str:
                    continue
                
                # Try common date formats with 4-digit years first
                parsed_date = None
                for fmt in ["%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d", "%d/%m/%Y"]:
                    try:
                        parsed_date = datetime.strptime(date_str, fmt).date()
                        break
                    except ValueError:
                        continue
                
                # Try 2-digit year formats - handle intelligently
                if not parsed_date:
                    for fmt in ["%m/%d/%y", "%d/%m/%y", "%m-%d-%y"]:
                        try:
                            temp_date = datetime.strptime(date_str, fmt)
                            year = temp_date.year
                            
                            # Python's %y: 00-68 -> 2000-2068, 69-99 -> 1969-1999
                            # For operational costs in 2025, we want dates in 2020s
                            # If parsed year is before 2020, assume it should be in 2000s
                            if year < 2020:
                                # Extract the last 2 digits and assume 2000s
                                last_two_digits = year % 100
                                # If last_two_digits is 0-30, it's likely 2000-2030
                                # If it's 31-99, it was parsed as 1931-1999, but should be 2031-2099
                                # For now, assume all should be 2000s (20xx)
                                year = 2000 + last_two_digits
                            
                            parsed_date = temp_date.replace(year=year).date()
                            break
                        except (ValueError, TypeError):
                            continue
                
                if parsed_date:
                    # Validate date is reasonable (between 2020 and 2030 for operational costs)
                    if 2020 <= parsed_date.year <= 2030:
                        date_column_indices.append((idx, parsed_date))
                    else:
                        logger.debug(f"Skipping date {parsed_date} (year {parsed_date.year} outside expected range)")
            except Exception:
                continue
        
        # Sort date columns by date
        date_column_indices.sort(key=lambda x: x[1])
        
        # Find the cost category rows we're looking for
        cost_categories = {
            "warehouse": ["warehouse costs per kg", "warehouse cost per kg", "warehouse costs"],
            "fulfilment": ["fulfilment costs per kg", "fulfillment costs per kg", "fulfilment cost per kg", "fulfillment cost per kg", "fulfilment costs"],
            "last_mile": ["last mile costs per kg", "last mile cost per kg", "last mile costs", "last-mile costs per kg"]
        }
        
        # Find "Day" column index
        day_col_idx = None
        for idx, header in enumerate(headers):
            if header.lower().strip() == "day":
                day_col_idx = idx
                break
        
        if day_col_idx is None:
            logger.warning("Could not find 'Day' column in daily operational costs sheet")
            return []
        
        # Known row numbers for cost categories (sheet row numbers, 1-indexed)
        # Row 51: Warehouse costs per Kg
        # Row 58: Fulfilment costs per Kg
        # Row 71: Last Mile costs per Kg
        # Convert to 0-indexed for all_values array (row 1 = index 0, row 51 = index 50)
        known_cost_rows = {
            50: "warehouse",  # Row 51 in sheet (0-indexed: 50)
            57: "fulfilment",  # Row 58 in sheet (0-indexed: 57)
            70: "last_mile",   # Row 71 in sheet (0-indexed: 70)
        }
        
        # Process data rows
        daily_costs: list[dict[str, Any]] = []
        
        # Process all rows starting from row 2 (index 1 in all_values)
        for row_index_in_array, row in enumerate(all_values[1:], start=1):  # Start from index 1 (row 2 in sheet)
            # row_index_in_array is the index in all_values (0-indexed, starting from 1 for row 2)
            # Actual sheet row number = row_index_in_array + 1
            
            if len(row) <= day_col_idx:
                continue
            
            day_value = str(row[day_col_idx]).strip().lower()
            if not day_value:
                continue
            
            # Check if this is a known cost category row by row number
            category_type = None
            category_name = None
            
            # Check if this row index matches one of the known cost category rows
            if row_index_in_array in known_cost_rows:
                category_type = known_cost_rows[row_index_in_array]
                category_name = day_value.title()
                logger.debug(f"Found {category_type} costs at row {row_index_in_array + 1} (index {row_index_in_array})")
            else:
                # Fallback: Check if this row matches any cost category by keywords
                for cat_type, keywords in cost_categories.items():
                    if any(keyword in day_value for keyword in keywords):
                        category_type = cat_type
                        category_name = day_value.title()
                        logger.debug(f"Found {category_type} costs by keyword matching at row {row_index_in_array + 1}")
                        break
            
            if not category_type:
                continue
            
            # Extract cost values for each date column
            for col_idx, date_obj in date_column_indices:
                if col_idx < len(row):
                    cost_value = _parse_float(row[col_idx])
                    if cost_value > 0:  # Only include non-zero costs
                        daily_costs.append({
                            "date": date_obj.isoformat(),
                            "category": category_type,
                            "category_name": category_name,
                            "cost_per_kg": cost_value
                        })
        
        # Sort by date, then by category
        daily_costs.sort(key=lambda x: (x["date"], x["category"]))
        
        logger.info(f"Loaded {len(daily_costs)} daily operational cost records")
        return daily_costs
        
    except GoogleSheetsNotConfigured:
        logger.warning("Google Sheets not configured for daily operational costs")
        return []
    except Exception as exc:
        logger.error(f"Failed to load daily operational costs: {exc}")
        import traceback
        logger.debug(traceback.format_exc())
        return []


def load_product_metrics_data(return_window: bool = False) -> Union[List[Dict[str, Any]], Tuple[List[Dict[str, Any]], Optional[Dict[str, Any]]]]:
    """Aggregate product-level sales metrics from Google Sheets (primary) or ClickHouse/CSV fallback."""
    metrics: list[dict[str, Any]] = []
    
    # 1. Fetch Primary Data from Google Sheets
    sheet_result = fetch_sheet_metrics()
    sheet_metrics = sheet_result.get("metrics", []) if sheet_result else []
    sheet_window = sheet_result.get("window") if sheet_result else None
    
    # Normalize and aggregate sheet metrics to handle aliases (e.g. Fossolia -> Green Beans)
    canonical_map: dict[str, dict[str, Any]] = {}
    
    for m in sheet_metrics:
        raw_name = str(m.get("product_name", "")).strip()
        if not raw_name:
            continue
        canonical = _normalize_product_name(raw_name)
        
        if canonical not in canonical_map:
            # Clone and set canonical name
            entry = m.copy()
            entry["product_name"] = canonical
            canonical_map[canonical] = entry
        else:
            # Aggregate duplicate canonical entries
            entry = canonical_map[canonical]
            entry["total_volume_kg"] += m.get("total_volume_kg", 0.0)
            entry["sgl_volume_kg"] += m.get("sgl_volume_kg", 0.0)
            entry["normal_volume_kg"] += m.get("normal_volume_kg", 0.0)
            entry["total_revenue_etb"] += m.get("total_revenue_etb", 0.0)
            entry["total_cost_etb"] = (entry.get("total_cost_etb") or 0.0) + (m.get("total_cost_etb") or 0.0)
            entry["order_count"] += m.get("order_count", 0)
            # Recalculate avg selling price
            vol = entry["total_volume_kg"]
            entry["avg_selling_price"] = entry["total_revenue_etb"] / vol if vol > 0 else 0.0
            # Recalculate gross profit
            entry["gross_profit_etb"] = entry["total_revenue_etb"] - (entry.get("total_cost_etb") or 0.0)
    
    metrics = list(canonical_map.values())
    existing_products = {c.lower() for c in canonical_map.keys()}

    week_start, week_end, week_start_dt, week_end_dt = _get_week_window_datetimes()
    
    # Override with sheet window if available to ensure consistency across all sources
    if sheet_window:
        week_start = sheet_window["start"]
        week_end = sheet_window["end"]
        week_start_dt = datetime.combine(week_start, datetime.min.time())
        week_end_dt = datetime.combine(week_end + timedelta(days=1), datetime.min.time())

    week_start_str = week_start_dt.strftime("%Y-%m-%d %H:%M:%S")
    week_end_str = week_end_dt.strftime("%Y-%m-%d %H:%M:%S")
    sales_summary = _get_sales_purchase_summary_for_window(week_start.isoformat(), week_end.isoformat())

    client = get_clickhouse_client()
    if not client:
        logger.warning("ClickHouse unavailable for product metrics")
    
    if client:
        try:
            query = f"""
        SELECT
            any(
                coalesce(
                    pn.name,
                    toString(gd.product_id)
                )
            ) AS product_name,
            gd.product_id AS product_id,
            sum(gc.quantity) AS total_volume_kg,
            sumIf(
                gc.quantity,
                gd.deal_type IN (
                    'SUPER_GROUP',
                    'SUPER_GROUP_FLASH_SALE',
                    'SUPER_GROUP_REGULAR',
                    'SUPER_GROUP_RECURRENT'
                )
            ) AS sgl_volume_kg,
            sum(gc.quantity)
            - sumIf(
                gc.quantity,
                gd.deal_type IN (
                    'SUPER_GROUP',
                    'SUPER_GROUP_FLASH_SALE',
                    'SUPER_GROUP_REGULAR',
                    'SUPER_GROUP_RECURRENT'
                )
            ) AS normal_volume_kg,
            round(avgIf(gd.group_price, gd.group_price > 0), 2) AS avg_selling_price,
            argMax(gd.group_price, o.created_at) AS latest_selling_price,
            sum(gc.quantity * gd.group_price) AS total_revenue_etb,
            countDistinct(o.id) AS order_count
        FROM orders AS o
        JOIN groups_carts AS gc ON o.groups_carts_id = gc.id
        JOIN groups AS g ON gc.group_id = g.id
        JOIN group_deals AS gd ON g.group_deals_id = gd.id
        LEFT JOIN products AS p ON gd.product_id = p.id
        LEFT JOIN product_names AS pn ON pn.id = p.name_id
        WHERE
            o._peerdb_is_deleted = 0
            AND gc._peerdb_is_deleted = 0
            AND g._peerdb_is_deleted = 0
            AND gd._peerdb_is_deleted = 0
            AND o.status = 'COMPLETED' AND o.deleted_at IS NULL
            AND gc.status = 'COMPLETED' AND gc.deleted_at IS NULL
            AND g.status = 'COMPLETED' AND g.deleted_at IS NULL
            AND o.created_at >= toDateTime('{week_start_str}')
            AND o.created_at < toDateTime('{week_end_str}')
        GROUP BY gd.product_id
        HAVING total_volume_kg > 0
        ORDER BY total_volume_kg DESC
        """
            result = client.query(query)
            for row in result.named_results():
                product_name = row.get("product_name")
                if not product_name:
                    continue
                
                # Skip if already fetched from Sheet
                canonical = _normalize_product_name(str(product_name)).lower()
                if canonical in existing_products:
                    continue

                entry = {
                    "product_name": str(product_name),
                    "product_id": row.get("product_id"),
                    "total_volume_kg": float(row.get("total_volume_kg", 0.0)),
                    "sgl_volume_kg": float(row.get("sgl_volume_kg", 0.0)),
                    "normal_volume_kg": float(row.get("normal_volume_kg", 0.0)),
                    "avg_selling_price": _parse_float(row.get("avg_selling_price")),
                    "latest_selling_price": _parse_float(row.get("latest_selling_price")),
                    "total_revenue_etb": _parse_float(row.get("total_revenue_etb")),
                    "order_count": int(row.get("order_count", 0)),
                }
                _sanitize_numeric_fields(entry)
                metrics.append(entry)
            
            if metrics:
                _apply_volume_ratio_overrides(metrics)
                if return_window:
                    return metrics, sheet_window
                return metrics
        except Exception as exc:  # pylint: disable=broad-except
            logger.error("Failed to load product metrics from ClickHouse: %s", exc)

    # Fallback to CSV data (SGL orders) if ClickHouse unavailable and not in Sheet
    # NOTE: This fallback is disabled because it lacks date filtering and causes massive inflation
    # by summing all-time history. We rely on Google Sheets or ClickHouse.
    if False and not SGL_ORDER_PRICE_CSV.exists():
        if metrics:
             _apply_volume_ratio_overrides(metrics)
        if return_window:
            return metrics, sheet_window
        return metrics

    if False: # Disabled CSV fallback
        try:
            aggregates: dict[str, dict[str, Any]] = {}
            with SGL_ORDER_PRICE_CSV.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    product = (row.get("product_name") or "").strip()
                    if not product:
                        continue
                    
                    # Skip if already present
                    canonical = _normalize_product_name(product).lower()
                    if canonical in existing_products:
                         continue

                    qty = _parse_float(row.get("total_group_kg") or row.get("leader_quantity_kg"))
                    price = _parse_float(row.get("group_price"))
                    deal_type = (row.get("group_deal_type") or "").strip().upper()

                    entry = aggregates.setdefault(
                        product,
                        {
                            "product_name": product,
                            "product_id": None,
                            "total_volume_kg": 0.0,
                            "sgl_volume_kg": 0.0,
                            "normal_volume_kg": 0.0,
                            "avg_selling_price": 0.0,
                            "latest_selling_price": 0.0,
                            "total_revenue_etb": 0.0,
                            "order_count": 0,
                        },
                    )
                    entry["total_volume_kg"] += qty
                    if deal_type in {"SUPER_GROUP", "SUPER_GROUP_FLASH_SALE"}:
                        entry["sgl_volume_kg"] += qty
                    else:
                        entry["normal_volume_kg"] += qty
                    entry["total_revenue_etb"] += qty * price
                    entry["order_count"] += 1
                    entry["latest_selling_price"] = price

            for entry in aggregates.values():
                volume = entry["total_volume_kg"]
                entry["avg_selling_price"] = (
                    entry["total_revenue_etb"] / volume if volume else 0.0
                )
                _sanitize_numeric_fields(entry)
                metrics.append(entry)
                # Add to existing set so we don't double add in next fallback step
                existing_products.add(_normalize_product_name(entry["product_name"]).lower())

        except Exception as exc:  # pylint: disable=broad-except
            logger.error("Failed to load product metrics from CSV fallback: %s", exc)

    _apply_volume_ratio_overrides(metrics)

    # Check sales summary for anything else
    for canonical, info in sales_summary.items():
        key = canonical.lower()
        if key in existing_products:
            continue
        weekly_qty = info.get("weekly_quantity") or 0.0
        avg_sell = info.get("avg_selling_price") or info.get("latest_selling_price") or 0.0
        entry = {
            "product_name": canonical,
            "product_id": None,
            "total_volume_kg": weekly_qty,
            "sgl_volume_kg": weekly_qty,
            "normal_volume_kg": 0.0,
            "avg_selling_price": avg_sell,
            "latest_selling_price": info.get("latest_selling_price") or avg_sell,
            "total_revenue_etb": info.get("weekly_sales_revenue") or 0.0,
            "order_count": int(round(weekly_qty)),
        }
        _sanitize_numeric_fields(entry)
        metrics.append(entry)

    if return_window:
        return metrics, sheet_window
    return metrics

@app.get("/")
async def root():
    """Serve frontend index.html or return API message if frontend not available."""
    if FRONTEND_DIST.exists() and (FRONTEND_DIST / "index.html").exists():
        return FileResponse(FRONTEND_DIST / "index.html")
    return {"message": "Delivery Map Analytics API is running", "status": "healthy"}

@app.get("/api/health")
async def health_check():
    """Detailed health check with database connectivity."""
    try:
        client = get_clickhouse_client()
        if client:
            result = client.query("SELECT 1 as test")
            return {
                "status": "healthy",
                "database": "connected",
                "message": "API and database are operational"
            }
        else:
            # API is healthy even if database is disconnected
            return {
                "status": "healthy",
                "database": "disconnected",
                "message": "API is operational (database unavailable)"
            }
    except Exception as e:
        # API is still healthy even if database check fails
        return {
            "status": "healthy",
            "database": "disconnected",
            "message": "API is operational (database unavailable)",
            "error": str(e)
        }

@app.get("/api/data")
def get_delivery_data():
    """Fetch delivery data from ClickHouse."""
    client = get_clickhouse_client()
    
    # Get week window even if database is unavailable
    week_start, week_end, week_start_dt, week_end_dt = _get_week_window_datetimes()
    
    # Override with sheet window if available to ensure consistency across all sources
    sheet_result = fetch_sheet_metrics()
    if sheet_result and sheet_result.get("window"):
        week_start = sheet_result["window"]["start"]
        week_end = sheet_result["window"]["end"]
        week_start_dt = datetime.combine(week_start, datetime.min.time())
        week_end_dt = datetime.combine(week_end + timedelta(days=1), datetime.min.time())
    
    if not client:
        logger.warning("ClickHouse unavailable for delivery data")
        return {
            "records": [],
            "count": 0,
            "window": {
                "start": week_start.isoformat(),
                "end": week_end.isoformat(),
            },
        }
    
    try:
        
        # Override with sheet window if available to ensure consistency across all sources
        sheet_result = fetch_sheet_metrics()
        if sheet_result and sheet_result.get("window"):
            week_start = sheet_result["window"]["start"]
            week_end = sheet_result["window"]["end"]
            week_start_dt = datetime.combine(week_start, datetime.min.time())
            week_end_dt = datetime.combine(week_end + timedelta(days=1), datetime.min.time())

        start_str = week_start_dt.strftime("%Y-%m-%d %H:%M:%S")
        end_str = week_end_dt.strftime("%Y-%m-%d %H:%M:%S")
        
        # Your SQL query
        query = f"""
        SELECT
            /* Existing bucket */
            CASE
                WHEN gd.deal_type IN ('SUPER_GROUP', 'SUPER_GROUP_FLASH_SALE') THEN 'SUPER_GROUPS'
                WHEN gd.deal_type IN ('NORMAL', 'FLASH_SALE') THEN 'NORMAL_GROUPS'
                ELSE gd.deal_type
            END AS group_deal_category,

            /* Leader identity */
            g.created_by                                   AS group_created_by,
            any(u.name)                                    AS leader_name,
            any(u.phone)                                   AS leader_phone,

            /* Delivery info */
            dl.location                                    AS delivery_coordinates,
            dl.name                                        AS delivery_location_name,

            /* New metrics you asked for */
            sum(gc.quantity)                               AS total_kg,                       -- total KG in the period
            toDate(max(o.created_at))                      AS last_order_date,                -- last order date
            countDistinct(gd.product_id)                   AS products_ordered,               -- number of distinct products ordered
            sum(gc.quantity) / nullIf(countDistinct(toDate(o.created_at)), 0) 
                                                           AS avg_kg_per_ordering_day,        -- average KG per active ordering day

            /* Existing metrics */
            countDistinct(gc.user_id)                      AS unique_group_members,
            countDistinct(g.id)                            AS total_groups,

            countIf(toDayOfWeek(o.created_at) = 1)         AS monday_orders,
            countIf(toDayOfWeek(o.created_at) = 2)         AS tuesday_orders,
            countIf(toDayOfWeek(o.created_at) = 3)         AS wednesday_orders,
            countIf(toDayOfWeek(o.created_at) = 4)         AS thursday_orders,
            countIf(toDayOfWeek(o.created_at) = 5)         AS friday_orders,
            countIf(toDayOfWeek(o.created_at) = 6)         AS saturday_orders,
            countIf(toDayOfWeek(o.created_at) = 7)         AS sunday_orders,

            countDistinct(toDate(o.created_at))            AS active_days,
            countDistinct(o.id)                            AS total_orders

        FROM orders AS o
        JOIN groups_carts AS gc
          ON o.groups_carts_id = gc.id
        JOIN groups AS g
          ON gc.group_id = g.id
        JOIN group_deals AS gd
          ON g.group_deals_id = gd.id
        JOIN delivery_location AS dl
          ON o.location_id = dl.id
        JOIN users u
          ON u.id = g.created_by
        WHERE
          o._peerdb_is_deleted = 0
          AND gc._peerdb_is_deleted = 0
          AND g._peerdb_is_deleted = 0
          AND gd._peerdb_is_deleted = 0
          /* (Optional but recommended) Keep only completed, non-deleted facts */
          AND o.status = 'COMPLETED'      AND o.deleted_at IS NULL
          AND gc.status = 'COMPLETED'     AND gc.deleted_at IS NULL
          AND g.status = 'COMPLETED'      AND g.deleted_at IS NULL
          /* Date range */
          AND o.created_at >= toDateTime('{start_str}')
          AND o.created_at <  toDateTime('{end_str}')
        GROUP BY
          group_deal_category,
          g.created_by,
          dl.location,
          dl.name
        ORDER BY
          group_deal_category ASC,
          unique_group_members DESC
        """
        
        result = client.query(query)
        
        # Convert to list of dictionaries
        data = []
        for row in result.result_rows:
            # Parse coordinates
            lat, lon = parse_coordinates(row[4])  # delivery_coordinates
            
            data.append({
                "group_deal_category": row[0],
                "group_created_by": row[1],
                "leader_name": row[2],
                "leader_phone": row[3],
                "delivery_coordinates": row[4],
                "delivery_location_name": row[5],
                "total_kg": float(row[6]) if row[6] is not None else 0.0,
                "last_order_date": str(row[7]) if row[7] is not None else None,
                "products_ordered": int(row[8]),
                "avg_kg_per_ordering_day": float(row[9]) if row[9] is not None else 0.0,
                "unique_group_members": int(row[10]),
                "total_groups": int(row[11]),
                "monday_orders": int(row[12]),
                "tuesday_orders": int(row[13]),
                "wednesday_orders": int(row[14]),
                "thursday_orders": int(row[15]),
                "friday_orders": int(row[16]),
                "saturday_orders": int(row[17]),
                "sunday_orders": int(row[18]),
                "active_days": int(row[19]),
                "total_orders": int(row[20]),
                "latitude": lat,
                "longitude": lon
            })
        
        return {
            "records": data,
            "count": len(data),
            "window": {
                "start": week_start.isoformat(),
                "end": week_end.isoformat(),
            },
        }
        
    except Exception as e:
        logger.error(f"Error fetching data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch data: {str(e)}")

@app.get("/api/statistics")
def get_statistics():
    """Get aggregated statistics."""
    client = get_clickhouse_client()
    if not client:
        logger.warning("ClickHouse unavailable for statistics")
        return {
            "totalRecords": 0,
            "normalGroups": 0,
            "superGroups": 0,
            "totalOrders": 0,
            "avgOrdersPerGroup": 0.0,
            "maxOrders": 0,
            "uniqueLocations": 0,
            "avgMembersPerGroup": 0.0,
            "geographicBounds": {
                "minLat": 0.0,
                "maxLat": 0.0,
                "minLon": 0.0,
                "maxLon": 0.0
            }
        }
    
    try:
        week_start, week_end, week_start_dt, week_end_dt = _get_week_window_datetimes()
        
        # Override with sheet window if available to ensure consistency across all sources
        sheet_result = fetch_sheet_metrics()
        if sheet_result and sheet_result.get("window"):
            week_start = sheet_result["window"]["start"]
            week_end = sheet_result["window"]["end"]
            week_start_dt = datetime.combine(week_start, datetime.min.time())
            week_end_dt = datetime.combine(week_end + timedelta(days=1), datetime.min.time())

        start_str = week_start_dt.strftime("%Y-%m-%d %H:%M:%S")
        end_str = week_end_dt.strftime("%Y-%m-%d %H:%M:%S")
        
        # Simple statistics query - just return basic counts
        stats_query = f"""
        SELECT
            COUNT(DISTINCT o.id) AS total_orders,
            COUNT(DISTINCT CASE 
                WHEN gd.deal_type IN ('NORMAL', 'FLASH_SALE') THEN o.id 
            END) AS normal_group_orders,
            COUNT(DISTINCT CASE 
                WHEN gd.deal_type IN ('SUPER_GROUP', 'SUPER_GROUP_FLASH_SALE') THEN o.id 
            END) AS super_group_orders,
            COUNT(DISTINCT dl.id) AS unique_locations
        FROM orders AS o
        JOIN groups_carts AS gc ON o.groups_carts_id = gc.id
        JOIN groups AS g ON gc.group_id = g.id
        JOIN group_deals AS gd ON g.group_deals_id = gd.id
        JOIN delivery_location AS dl ON o.location_id = dl.id
        WHERE o._peerdb_is_deleted = 0
          AND gc._peerdb_is_deleted = 0
          AND g._peerdb_is_deleted = 0
          AND gd._peerdb_is_deleted = 0
          AND o.status = 'COMPLETED' AND o.deleted_at IS NULL
          AND gc.status = 'COMPLETED' AND gc.deleted_at IS NULL
          AND g.status = 'COMPLETED' AND g.deleted_at IS NULL
          AND o.created_at >= toDateTime('{start_str}')
          AND o.created_at <  toDateTime('{end_str}')
        """
        
        result = client.query(stats_query)
        row = result.result_rows[0]
        
        total_orders = int(row[0])
        normal_orders = int(row[1])
        super_orders = int(row[2])
        unique_locs = int(row[3])
        
        return {
            "totalRecords": unique_locs,
            "normalGroups": normal_orders,
            "superGroups": super_orders,
            "totalOrders": total_orders,
            "avgOrdersPerGroup": 10.0,
            "maxOrders": 1000,
            "uniqueLocations": unique_locs,
            "avgMembersPerGroup": 10.0,
            "geographicBounds": {
                "minLat": 8.0,
                "maxLat": 10.0,
                "minLon": 38.0,
                "maxLon": 39.0
            },
            "window": {
                "start": week_start.isoformat(),
                "end": week_end.isoformat(),
            },
        }
        
    except Exception as e:
        logger.error(f"Error fetching statistics: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch statistics: {str(e)}")

@app.get("/api/forecast/weekly-summary")
def forecast_weekly_summary():
    """
    Provide the baseline weekly orders and volume using the latest completed ClickHouse data.
    Returns null metrics when unavailable so the frontend can fall back gracefully.
    """
    summary = _fetch_latest_weekly_summary()
    if summary:
        return summary
    return {
        "week_start": None,
        "orders": None,
        "total_kg": None,
        "source": "fallback",
    }

@app.get("/api/forecast/products")
def forecast_products():
    """Return distinct benchmark product names from local shop CSV (agri focus)."""
    products = load_local_shop_products()
    return {"products": products}

@app.get("/api/forecast/shares")
def forecast_product_shares():
    """Return global demand shares for products restricted to benchmark list."""
    products = set(load_local_shop_products())
    result = compute_product_shares(products)
    return result


@app.get("/api/benchmark/local-prices")
def get_local_shop_prices():
    """Expose latest local shop price per product."""
    prices = load_local_shop_price_map()
    return {"prices": list(prices.values())}

@app.get("/api/benchmark/locations")
def get_benchmark_locations():
    """Return unique benchmark locations with coordinates and category."""
    if not BENCHMARK_API_URL or not BENCHMARK_API_KEY:
        return {"locations": []}
    
    from services.geocoding import geocode_location
    
    today = datetime.utcnow().date()
    date_from = today - timedelta(days=6)  # 7-day window
    
    params = {
        "dateFrom": date_from.strftime("%Y-%m-%d"),
        "dateTo": today.strftime("%Y-%m-%d"),
        "frequency": "daily",
        "comparisonType": "avg",
        "locationGroups": ",".join(BENCHMARK_LOCATION_GROUPS),
    }
    
    headers = {
        "apikey": BENCHMARK_API_KEY,
        "Authorization": f"Bearer {BENCHMARK_API_KEY}",
        "Content-Type": "application/json",
    }
    
    try:
        with httpx.Client(timeout=15.0) as client:
            response = client.get(BENCHMARK_API_URL, params=params, headers=headers)
            response.raise_for_status()
            payload = response.json()
    except Exception as e:
        logger.warning("Failed to fetch benchmark locations: %s", e)
        return {"locations": []}
    
    data_entries = payload.get("data") or []
    if not data_entries:
        return {"locations": []}
    
    # Track unique locations by (location_name, location_group)
    unique_locations: Dict[Tuple[str, str], Dict[str, Any]] = {}
    
    for entry in data_entries:
        location_name = (entry.get("location") or "").strip()
        location_group = (entry.get("location_group") or "").strip().lower()
        
        if not location_name or not location_group:
            continue
        
        key = (location_name, location_group)
        
        # Initialize location if not exists
        if key not in unique_locations:
            # Try to get coordinates from API
            lat_raw = entry.get("latitude") or entry.get("lat") or entry.get("location_latitude")
            lon_raw = entry.get("longitude") or entry.get("lon") or entry.get("location_longitude")
            
            lat_val = None
            lon_val = None
            
            try:
                if lat_raw is not None and lon_raw is not None:
                    lat_val = float(lat_raw)
                    lon_val = float(lon_raw)
            except (TypeError, ValueError):
                pass
            
            # If no coordinates, try geocoding
            if (lat_val is None or lon_val is None) and location_name:
                coords = geocode_location(location_name, location_group)
                if coords:
                    lat_val, lon_val = coords
            
            # Only include if we have coordinates
            if lat_val is not None and lon_val is not None:
                unique_locations[key] = {
                    "location": location_name,
                    "location_group": location_group,
                    "latitude": lat_val,
                    "longitude": lon_val,
                    "products": []
                }
        
        # Add product info if location exists
        if key in unique_locations:
            product_name = (entry.get("product_name") or entry.get("product") or "").strip()
            price = _parse_float(entry.get("price"))
            if product_name and price > 0:
                # Check if product already exists to avoid duplicates (e.g. multiple entries for same date/product)
                existing_products = unique_locations[key]["products"]
                # Simple check: if product name exists, maybe update price if newer?
                # For now, we'll just append and let frontend filter or show latest. 
                # Actually, let's keep the latest price for each product per location.
                
                existing_idx = next((i for i, p in enumerate(existing_products) if p["name"] == product_name), -1)
                
                new_product_entry = {
                    "name": product_name,
                    "price": price,
                    "date": entry.get("date")
                }
                
                if existing_idx >= 0:
                    # Compare dates to keep latest
                    current = existing_products[existing_idx]
                    current_date = _parse_datetime(current.get("date"))
                    new_date = _parse_datetime(new_product_entry.get("date"))
                    
                    if new_date and (not current_date or new_date > current_date):
                        existing_products[existing_idx] = new_product_entry
                else:
                    existing_products.append(new_product_entry)
    
    return {"locations": list(unique_locations.values())}

@app.get("/api/forecast/elasticities")
def forecast_product_elasticities():
    """Return measured product-specific elasticities from analysis."""
    elasticity_file = PROJECT_ROOT / "data_points" / "ANALYSIS_product_elasticity.csv"
    try:
        import csv
        elasticities = {}
        with elasticity_file.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                product = row.get("product", "").strip()
                elasticity_str = row.get("elasticity", "").strip()
                if product and elasticity_str:
                    try:
                        elasticities[product] = float(elasticity_str)
                    except Exception:
                        pass
        return {"elasticities": elasticities}
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Elasticity analysis not found. Run analyze_elasticity_personas.py first.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to load elasticities: {e}")

@app.get("/api/forecast/personas")
def forecast_personas():
    """Return persona definitions and summary stats."""
    persona_file = PROJECT_ROOT / "data_points" / "ANALYSIS_persona_summary.csv"
    try:
        import csv
        personas = []
        with persona_file.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                personas.append({
                    "name": row.get("Persona", ""),
                    "elasticity": float(row.get("Assigned Elasticity", -1.0)),
                    "leader_count": int(row.get("Leader Count", 0)),
                    "avg_kg_per_day": float(row.get("Avg KG per Day", 0)),
                    "avg_price_vs_local": float(row.get("Avg Price vs Local (%)", 0))
                })
        return {"personas": personas}
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Persona analysis not found. Run analyze_elasticity_personas.py first.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to load personas: {e}")

@app.get("/api/forecast/commissions")
def forecast_commissions():
    """Return commission recommendations from COMMISSION_LOOKUP.csv."""
    if not COMMISSION_LOOKUP_CSV.exists():
        raise HTTPException(status_code=404, detail="Commission lookup not found.")
    commissions = {}
    try:
        with COMMISSION_LOOKUP_CSV.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                product = row.get("product_name", "").strip()
                if product:
                    commissions[product] = {
                        "recommended_commission": float(row.get("recommended_commission_etb", "3.0")),
                        "commission_pct_of_price": float(row.get("commission_as_pct_of_price", "10.0")),
                        "min_commission": float(row.get("minimum_commission", "1.0")),
                        "max_commission": float(row.get("maximum_commission", "10.0")),
                        "notes": row.get("notes", "")
                    }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed reading commission CSV: {e}")
    return {"commissions": commissions}


def _parse_iso_date(value: Optional[str], default: date) -> date:
    if value is None:
        return default
    if isinstance(value, str):
        try:
            return datetime.strptime(value.strip(), "%Y-%m-%d").date()
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=f"Invalid date format: {value}. Use YYYY-MM-DD.") from exc
    # FastAPI Query default objects (used when invoking directly) are not strings; fall back to default.
    return default


@app.get("/api/personas/leaders")
def get_persona_leaders(
    start_date: Optional[str] = Query(
        None, description="Inclusive start date (YYYY-MM-DD) for sensitivity analysis"
    ),
    end_date: Optional[str] = Query(
        None, description="Inclusive end date (YYYY-MM-DD) for sensitivity analysis"
    ),
):
    """
    Return deduplicated SGL leaders with persona, coordinates, and price sensitivity metric.
    """
    leaders_file = PROJECT_ROOT / "data_points" / "SGL_persona_leaders_unique.csv"
    if not leaders_file.exists():
        raise HTTPException(status_code=404, detail="SGL persona leaders dataset not found.")

    default_start = date(2025, 4, 1)
    default_end = date(2025, 11, 10)
    analysis_start = _parse_iso_date(start_date, default_start)
    analysis_end = _parse_iso_date(end_date, default_end)
    if analysis_end < analysis_start:
        raise HTTPException(status_code=400, detail="end_date must be on or after start_date")

    sensitivity_by_phone: Dict[str, Dict[str, Any]] = {}
    sensitivity_by_leader_id: Dict[str, Dict[str, Any]] = {}
    sensitivity_by_name: Dict[str, Dict[str, Any]] = {}

    client = None
    leader_coords_map = _load_leader_coordinate_map()

    try:
        client = get_clickhouse_client()
        if client:
            sensitivity_maps = compute_leader_sensitivity(client, analysis_start, analysis_end, leader_coords_map)
            sensitivity_by_phone = sensitivity_maps.get("by_phone", {})
            sensitivity_by_leader_id = sensitivity_maps.get("by_leader_id", {})
            sensitivity_by_name = sensitivity_maps.get("by_name", {})
        else:
            logger.warning("ClickHouse client unavailable; continuing without sensitivity metrics")
    except HTTPException:
        # Propagate upstream errors (likely ClickHouse configuration issues)
        raise
    except Exception as exc:  # pylint: disable=broad-except
        logger.warning("Failed to compute sensitivity metrics; continuing without them: %s", exc)
    finally:
        if client is not None:
            try:
                client.close()  # type: ignore[attr-defined]
            except Exception:  # pylint: disable=broad-except
                pass

    try:
        import csv

        leaders: list[dict[str, Any]] = []
        with leaders_file.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                try:
                    lat = float(row.get("latest_delivery_latitude", "") or 0.0)
                    lon = float(row.get("latest_delivery_longitude", "") or 0.0)
                except ValueError:
                    continue

                if lat == 0.0 and lon == 0.0:
                    continue

                try:
                    price_sensitivity_val = float(row.get("price_sensitivity", "") or 0.0)
                except ValueError:
                    price_sensitivity_val = None

                leaders.append(
                    {
                        "phone": row.get("phone"),
                        "leader_name": row.get("leader_name"),
                        "persona": row.get("Persona"),
                        "total_kg_ordered": float(row.get("total_kg_ordered", 0) or 0),
                        "avg_kg_per_order_day": float(row.get("avg_kg_per_order_day", 0) or 0),
                        "wallet_commission": float(row.get("wallet_commission", 0) or 0),
                        "price_sensitivity": price_sensitivity_val,
                        "latitude": lat,
                        "longitude": lon,
                        "delivery_location": row.get("delivery_location"),
                    }
                )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to load SGL persona leaders: {exc}")

    period_payload = {"start_date": analysis_start.isoformat(), "end_date": analysis_end.isoformat()}

    enriched_leaders: list[dict[str, Any]] = []
    for leader in leaders:
        phone = (leader.get("phone") or "").strip() if leader.get("phone") else None
        leader_name = (leader.get("leader_name") or "").strip().lower() if leader.get("leader_name") else None

        metrics = None
        if phone:
            metrics = sensitivity_by_phone.get(phone)
        if metrics is None and leader_name:
            metrics = sensitivity_by_name.get(leader_name)
        if metrics is None:
            leader_id_candidate = leader.get("leader_phone") or leader.get("leader_id")
            if leader_id_candidate:
                metrics = sensitivity_by_leader_id.get(str(leader_id_candidate))

        if metrics:
            leader.update(
                {
                    "combined_sensitivity_etb": metrics.get("combined_sensitivity_etb"),
                    "local_discount_etb": metrics.get("local_discount_etb"),
                    "distribution_discount_etb": metrics.get("distribution_discount_etb"),
                    "local_discount_pct": metrics.get("local_discount_pct"),
                    "distribution_discount_pct": metrics.get("distribution_discount_pct"),
                    "sensitivity_coverage_days": metrics.get("coverage_days"),
                    "sensitivity_local_observations": metrics.get("local_observations"),
                    "sensitivity_distribution_observations": metrics.get("distribution_observations"),
                    "pct_volume_at_or_above_local": metrics.get("pct_volume_at_or_above_local"),
                    "sensitivity_source_flags": metrics.get("source_flags"),
                    "sensitivity_period": metrics.get("period", period_payload),
                    "sensitivity_total_kg": metrics.get("total_kg"),
                    "product_sensitivity": metrics.get("product_sensitivity", []),
                    # Backwards compatibility with legacy UI field
                    "price_sensitivity": metrics.get("combined_sensitivity_etb"),
                }
            )
        else:
            leader.update(
                {
                    "combined_sensitivity_etb": None,
                    "local_discount_etb": None,
                    "distribution_discount_etb": None,
                    "local_discount_pct": None,
                    "distribution_discount_pct": None,
                    "sensitivity_coverage_days": 0,
                    "sensitivity_local_observations": 0,
                    "sensitivity_distribution_observations": 0,
                    "pct_volume_at_or_above_local": None,
                    "sensitivity_source_flags": {"benchmark_api": False, "distribution_fallback": False},
                    "sensitivity_period": period_payload,
                    "sensitivity_total_kg": None,
                    "product_sensitivity": [],
                    "price_sensitivity": None,
                }
            )

        enriched_leaders.append(leader)

    return {"leaders": enriched_leaders}


@app.get("/api/sgl/retention")
def get_sgl_retention(
    start_date: Optional[str] = Query(
        None, description="Inclusive start date (YYYY-MM-DD) for retention analysis"
    ),
    end_date: Optional[str] = Query(
        None, description="Inclusive end date (YYYY-MM-DD) for retention analysis"
    ),
):
    default_start = date(2025, 4, 4)  # first Friday in April 2025
    _, current_week_end = _get_week_window()
    analysis_start = _parse_iso_date(start_date, default_start)
    analysis_end = _parse_iso_date(end_date, current_week_end)
    if analysis_end < analysis_start:
        raise HTTPException(status_code=400, detail="end_date must be on or after start_date")

    leader_coords_map = _load_leader_coordinate_map()

    client = get_clickhouse_client()
    if not client:
        logger.warning("ClickHouse unavailable for SGL retention")
        return {
            "products": [],
            "window": {
                "start": analysis_start.isoformat(),
                "end": analysis_end.isoformat(),
            },
            "message": "Database unavailable"
        }
    
    try:
        retention = compute_weekly_retention(client, analysis_start, analysis_end, leader_coords_map)
    except Exception as exc:  # pylint: disable=broad-except
        logger.error("Failed to compute SGL retention: %s", exc)
        raise HTTPException(status_code=500, detail="Failed to compute SGL retention data") from exc

    return {
        "products": retention,
        "window": {
            "start": analysis_start.isoformat(),
            "end": analysis_end.isoformat(),
        },
    }


# ---------- Cost Management Endpoints ----------


@app.get("/api/products/metrics")
def get_product_metrics():
    """Return product sales metrics for profitability and forecasting."""
    metrics, window = load_product_metrics_data(return_window=True)
    if window:
        start_str = window["start"].isoformat()
        end_str = window["end"].isoformat()
    else:
        week_start, week_end = _get_week_window()
        start_str = week_start.isoformat()
        end_str = week_end.isoformat()

    return {
        "metrics": metrics,
        "lookback_days": WEEK_LENGTH_DAYS,
        "window": {
            "start": start_str,
            "end": end_str,
        },
    }

@app.get("/api/costs/products")
def get_product_costs():
    """Return product cost structure from Google Sheets or CSV fallback."""
    products = load_product_costs_data()
    return {"products": products}

@app.get("/api/costs/operational")
def get_operational_costs():
    """Return operational cost breakdown from Google Sheets or CSV fallback."""
    costs = load_operational_costs_data()
    return {"costs": costs}

@app.get("/api/costs/daily-operational")
def get_daily_operational_costs(
    date_from: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="End date (YYYY-MM-DD)")
):
    """Return daily operational costs (Warehouse, Fulfilment, Last Mile) per kg from Google Sheet.
    
    Returns costs grouped by date and category. Optionally filter by date range.
    """
    from datetime import datetime
    
    daily_costs = load_daily_operational_costs_data()
    
    # Filter by date range if provided
    if date_from or date_to:
        filtered_costs = []
        for cost in daily_costs:
            cost_date = datetime.fromisoformat(cost["date"]).date()
            
            if date_from:
                from_date = datetime.fromisoformat(date_from).date()
                if cost_date < from_date:
                    continue
            
            if date_to:
                to_date = datetime.fromisoformat(date_to).date()
                if cost_date > to_date:
                    continue
            
            filtered_costs.append(cost)
        
        daily_costs = filtered_costs
    
    # Group by date for easier frontend consumption
    costs_by_date: dict[str, dict[str, Any]] = {}
    for cost in daily_costs:
        date_key = cost["date"]
        if date_key not in costs_by_date:
            costs_by_date[date_key] = {
                "date": date_key,
                "warehouse_cost_per_kg": None,
                "fulfilment_cost_per_kg": None,
                "last_mile_cost_per_kg": None
            }
        
        category = cost["category"]
        if category == "warehouse":
            costs_by_date[date_key]["warehouse_cost_per_kg"] = cost["cost_per_kg"]
        elif category == "fulfilment":
            costs_by_date[date_key]["fulfilment_cost_per_kg"] = cost["cost_per_kg"]
        elif category == "last_mile":
            costs_by_date[date_key]["last_mile_cost_per_kg"] = cost["cost_per_kg"]
    
    # Convert to list sorted by date
    result = sorted(costs_by_date.values(), key=lambda x: x["date"])
    
    return {
        "daily_costs": result,
        "count": len(result),
        "date_range": {
            "from": result[0]["date"] if result else None,
            "to": result[-1]["date"] if result else None
        } if result else None
    }

@app.get("/api/margins/daily-real")
def get_daily_real_margins(
    date_from: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="End date (YYYY-MM-DD)")
):
    """Calculate real margins by matching daily operational costs with daily volumes.
    
    Only includes data from 10/13/2025 onwards as specified.
    Returns daily margins with operational costs applied to volumes.
    """
    from datetime import datetime, date as date_type
    import pandas as pd
    
    # Minimum date: 10/13/2025
    MIN_DATE = date_type(2025, 10, 13)
    
    # Get daily operational costs
    daily_costs_raw = load_daily_operational_costs_data()
    
    # Get daily volumes from Google Sheet
    df = fetch_raw_sheet_data()
    if df is None or df.empty:
        return {
            "daily_margins": [],
            "count": 0,
            "error": "No volume data available"
        }
    
    # Filter by date range and minimum date
    if 'date_dt' in df.columns:
        df['date'] = df['date_dt'].dt.date
        if date_from:
            from_date = datetime.fromisoformat(date_from).date()
            df = df[df['date'] >= from_date]
        else:
            df = df[df['date'] >= MIN_DATE]
        
        if date_to:
            to_date = datetime.fromisoformat(date_to).date()
            df = df[df['date'] <= to_date]
    else:
        return {
            "daily_margins": [],
            "count": 0,
            "error": "Date column not found in volume data"
        }
    
    # Group volumes by date
    daily_volumes = df.groupby('date').agg({
        'final_volume_kg': 'sum',
        'final_revenue': 'sum',
        'final_cost': 'sum'
    }).reset_index()
    
    # Group operational costs by date
    costs_by_date: dict[str, dict[str, float]] = {}
    for cost in daily_costs_raw:
        cost_date_str = cost["date"]
        cost_date = datetime.fromisoformat(cost_date_str).date()
        
        # Only include dates from 10/13/2025 onwards
        if cost_date < MIN_DATE:
            continue
        
        if cost_date_str not in costs_by_date:
            costs_by_date[cost_date_str] = {
                "warehouse": 0.0,
                "fulfilment": 0.0,
                "last_mile": 0.0
            }
        
        category = cost["category"]
        if category in costs_by_date[cost_date_str]:
            costs_by_date[cost_date_str][category] = cost["cost_per_kg"]
    
    # Calculate margins for each day
    daily_margins = []
    for _, row in daily_volumes.iterrows():
        day_date = row['date']
        day_date_str = day_date.isoformat() if isinstance(day_date, date_type) else str(day_date)
        
        volume_kg = float(row['final_volume_kg'])
        revenue = float(row['final_revenue'])
        procurement_cost = float(row['final_cost'])
        
        # Get operational costs for this day
        day_costs = costs_by_date.get(day_date_str, {
            "warehouse": 0.0,
            "fulfilment": 0.0,
            "last_mile": 0.0
        })
        
        # Calculate total operational cost
        total_ops_cost_per_kg = (
            day_costs["warehouse"] + 
            day_costs["fulfilment"] + 
            day_costs["last_mile"]
        )
        
        total_ops_cost = total_ops_cost_per_kg * volume_kg if volume_kg > 0 else 0.0
        
        # Calculate real margin
        total_cost = procurement_cost + total_ops_cost
        margin = revenue - total_cost
        margin_pct = (margin / revenue * 100) if revenue > 0 else 0.0
        
        daily_margins.append({
            "date": day_date_str,
            "volume_kg": volume_kg,
            "revenue_etb": revenue,
            "procurement_cost_etb": procurement_cost,
            "warehouse_cost_etb": day_costs["warehouse"] * volume_kg if volume_kg > 0 else 0.0,
            "fulfilment_cost_etb": day_costs["fulfilment"] * volume_kg if volume_kg > 0 else 0.0,
            "last_mile_cost_etb": day_costs["last_mile"] * volume_kg if volume_kg > 0 else 0.0,
            "total_operational_cost_etb": total_ops_cost,
            "warehouse_cost_per_kg": day_costs["warehouse"],
            "fulfilment_cost_per_kg": day_costs["fulfilment"],
            "last_mile_cost_per_kg": day_costs["last_mile"],
            "total_operational_cost_per_kg": total_ops_cost_per_kg,
            "total_cost_etb": total_cost,
            "margin_etb": margin,
            "margin_pct": margin_pct
        })
    
    # Sort by date
    daily_margins.sort(key=lambda x: x["date"])
    
    return {
        "daily_margins": daily_margins,
        "count": len(daily_margins),
        "date_range": {
            "from": daily_margins[0]["date"] if daily_margins else None,
            "to": daily_margins[-1]["date"] if daily_margins else None
        } if daily_margins else None
    }

@app.get("/api/costs/operational/latest")
def get_latest_operational_costs():
    """Get the latest operational costs for use in playground.
    
    Returns the most recent values for warehouse, fulfilment, and last mile costs per kg.
    """
    from datetime import datetime
    
    daily_costs = load_daily_operational_costs_data()
    
    if not daily_costs:
        # Fallback to static operational costs
        static_costs = load_operational_costs_data()
        warehouse = 0.0
        fulfilment = 0.0
        last_mile = 0.0
        
        for cost in static_costs:
            cat = cost.get("cost_category", "").lower()
            if "warehouse" in cat:
                warehouse = cost.get("cost_per_kg", 0.0)
            elif "fulfilment" in cat:
                fulfilment = cost.get("cost_per_kg", 0.0)
            elif "last" in cat and "mile" in cat:
                last_mile = cost.get("cost_per_kg", 0.0)
        
        return {
            "warehouse_cost_per_kg": warehouse,
            "fulfilment_cost_per_kg": fulfilment,
            "last_mile_cost_per_kg": last_mile,
            "total_cost_per_kg": warehouse + fulfilment + last_mile,
            "source": "static",
            "date": None
        }
    
    # Find the latest date
    latest_date = max(datetime.fromisoformat(cost["date"]).date() for cost in daily_costs)
    latest_date_str = latest_date.isoformat()
    
    # Get costs for the latest date
    latest_costs = {
        "warehouse": 0.0,
        "fulfilment": 0.0,
        "last_mile": 0.0
    }
    
    for cost in daily_costs:
        if cost["date"] == latest_date_str:
            category = cost["category"]
            if category in latest_costs:
                latest_costs[category] = cost["cost_per_kg"]
    
    total = latest_costs["warehouse"] + latest_costs["fulfilment"] + latest_costs["last_mile"]
    
    return {
        "warehouse_cost_per_kg": latest_costs["warehouse"],
        "fulfilment_cost_per_kg": latest_costs["fulfilment"],
        "last_mile_cost_per_kg": latest_costs["last_mile"],
        "total_cost_per_kg": total,
        "source": "daily",
        "date": latest_date_str
    }

@app.get("/api/costs/tiers")
def get_sgl_tiers():
    """Return SGL tier definitions from SGL_TIERS.csv."""
    tiers_file = DATA_POINTS_DIR / "SGL_TIERS.csv"
    if not tiers_file.exists():
        raise HTTPException(status_code=404, detail="SGL tiers file not found.")
    
    tiers = []
    try:
        with tiers_file.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                tiers.append({
                    "tier_name": row.get("tier_name", "").strip(),
                    "description": row.get("description", "").strip(),
                    "commission_etb_per_kg": float(row.get("commission_etb_per_kg", "0")),
                    "cost_savings_logistics": float(row.get("cost_savings_logistics", "0")),
                    "cost_savings_packaging": float(row.get("cost_savings_packaging", "0")),
                    "cost_savings_assistant": float(row.get("cost_savings_assistant", "0"))
                })
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to load SGL tiers: {e}")
    
    return {"tiers": tiers}


# ========== B2B Analytics Endpoints ==========

@app.get("/api/b2b/business-overview")
async def get_b2b_business_overview(
    date_from: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="End date (YYYY-MM-DD)")
):
    """Get B2B business overview: revenue, orders, profit summary. Calculated from daily product orders."""
    try:
        from collections import defaultdict
        from datetime import datetime as dt
        
        client = get_b2b_mcp_client()
        purchase_price_service = get_b2b_purchase_price_service()
        date_range = format_date_range(date_from, date_to)
        
        # Get daily product orders data from Product Orders MCP
        daily_sales = await client.get_daily_sales_data(date_from, date_to)
        
        if not daily_sales or "items" not in daily_sales or len(daily_sales.get("items", [])) == 0:
            return {
                "revenue": 0,
                "orders": 0,
                "profit": 0,
                "profit_margin": 0,
                "total_customers": 0,
                "avg_order_value": 0,
                "date_range": date_range,
                "error": "No B2B data available for the specified date range"
            }
        
        # Calculate totals from daily data (already correctly calculated per day)
        total_revenue = 0.0
        total_orders = 0
        total_cogs = 0.0
        total_warehouse = 0.0
        total_delivery = 0.0
        unique_customers = set()
        
        for item in daily_sales["items"]:
            product_name = item.get("product_name")
            order_date_str = item.get("order_date") or item.get("sale_date")
            quantity_kg = float(item.get("quantity_kg") or item.get("quantity_sold") or 0.0)
            revenue = float(item.get("total_revenue") or item.get("revenue") or 0.0)
            order_count = int(item.get("number_of_orders") or item.get("order_count") or 0)
            
            # Include items with revenue > 0 even if quantity is 0 (some orders may have 0 quantity but still have revenue)
            # Or include items with orders > 0 (to count orders even if quantity/revenue is 0)
            if not product_name or (quantity_kg <= 0 and revenue <= 0 and order_count <= 0):
                continue
            
            # Parse order date
            try:
                if isinstance(order_date_str, str):
                    order_date = dt.strptime(order_date_str.split("T")[0], "%Y-%m-%d").date()
                else:
                    order_date = order_date_str
            except:
                continue
            
            # Get purchase price for this specific order date (with next-day offset)
            purchase_price, _ = purchase_price_service.get_purchase_price_for_sale_date(
                product_name, order_date
            )
            
            # Calculate daily COGS = purchase_price × quantity
            daily_cogs = purchase_price * quantity_kg if purchase_price > 0 else 0.0
            daily_warehouse = quantity_kg * 3.0  # 3 birr/kg
            daily_delivery = quantity_kg * 2.0   # 2 birr/kg
            
            # Sum totals
            total_revenue += revenue
            total_orders += order_count
            total_cogs += daily_cogs
            total_warehouse += daily_warehouse
            total_delivery += daily_delivery
            
            # Track unique customers (if customer_type_id is available)
            customer_type_id = item.get("customer_type_id")
            if customer_type_id:
                unique_customers.add(customer_type_id)
        
        # Calculate profit
        total_profit = total_revenue - total_cogs - total_warehouse - total_delivery
        profit_margin = (total_profit / total_revenue * 100) if total_revenue > 0 else 0.0
        avg_order_value = (total_revenue / total_orders) if total_orders > 0 else 0.0
        
        return {
            "revenue": round(total_revenue, 2),
            "orders": total_orders,
            "profit": round(total_profit, 2),
            "profit_margin": round(profit_margin, 2),
            "total_customers": len(unique_customers),
            "avg_order_value": round(avg_order_value, 2),
            "date_range": date_range,
            "costs": {
                "product_cogs": round(total_cogs, 2),
                "warehouse": round(total_warehouse, 2),
                "delivery": round(total_delivery, 2)
            }
        }
    except Exception as e:
        logger.error(f"Error fetching B2B business overview: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch B2B business overview: {str(e)}")


@app.get("/api/b2b/profit-margin-analysis")
async def get_b2b_profit_margin_analysis(
    date_from: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="End date (YYYY-MM-DD)")
):
    """Get B2B profit margin analysis: CM1, CM2, CM3 breakdown with actual purchase prices."""
    try:
        from datetime import datetime as dt
        
        client = get_b2b_mcp_client()
        purchase_price_service = get_b2b_purchase_price_service()
        date_range = format_date_range(date_from, date_to)
        
        # Get daily product orders data (new MCP provides this directly)
        daily_sales = await client.get_daily_sales_data(date_from, date_to)
        
        # Calculate profit margin analysis from daily data
        if not daily_sales or "items" not in daily_sales or len(daily_sales.get("items", [])) == 0:
            return {
                "margin_breakdown": {
                    "cm1_gross_profit": {"amount": 0, "percentage": "0%", "formula": "Revenue - Product COGS", "description": "Revenue minus direct product costs."},
                    "cm2_after_warehouse": {"amount": 0, "percentage": "0%", "formula": "CM1 - Warehouse Costs", "description": "CM1 minus warehouse handling costs."},
                    "cm3_net_profit": {"amount": 0, "percentage": "0%", "formula": "CM2 - Delivery Costs", "description": "CM2 minus last-mile delivery costs."},
                },
                "summary": {
                    "total_revenue": 0,
                    "product_cogs": 0,
                    "warehouse_costs": 0,
                    "delivery_costs": 0,
                    "final_profit": 0,
                    "final_margin": "0%",
                },
                "period": date_range,
                "error": "No B2B data available for the specified date range"
            }
        
        # Calculate from daily data (already fetched above)
        try:
            use_daily_data = daily_sales and "items" in daily_sales and len(daily_sales.get("items", [])) > 0
            
            if use_daily_data:
                # Calculate on daily basis: for each day, calculate COGS = purchase_price × quantity, then sum
                total_actual_cogs = 0.0
                total_actual_revenue = 0.0
                purchase_price_stats = defaultdict(int)
                
                for item in daily_sales["items"]:
                    product_name = item.get("product_name") or item.get("product_id", "")
                    if not product_name:
                        continue
                    
                    sale_date_str = item.get("sale_date") or item.get("order_date") or item.get("date")
                    quantity_kg = float(item.get("quantity_sold") or item.get("quantity_kg") or item.get("quantity") or item.get("weight_kg") or 0.0)
                    selling_price = float(item.get("selling_price") or item.get("price") or item.get("unit_price") or 0.0)
                    
                    if quantity_kg <= 0:
                        continue
                    
                    # Parse sale date
                    try:
                        if isinstance(sale_date_str, str):
                            sale_date = dt.strptime(sale_date_str.split("T")[0], "%Y-%m-%d").date()
                        else:
                            sale_date = sale_date_str
                    except:
                        continue
                    
                    # Calculate daily revenue - use total_revenue from MCP (already calculated correctly per day)
                    daily_revenue = float(item.get("total_revenue") or item.get("revenue") or 0.0)
                    # If revenue not provided, calculate from unit_price × quantity
                    if daily_revenue <= 0 and selling_price > 0:
                        daily_revenue = selling_price * quantity_kg
                    total_actual_revenue += daily_revenue
                    
                    # Get purchase price for this specific sale date (with next-day offset)
                    purchase_price, source = purchase_price_service.get_purchase_price_for_sale_date(
                        product_name, sale_date
                    )
                    
                    # Calculate daily COGS = purchase_price × quantity
                    if purchase_price > 0:
                        total_actual_cogs += purchase_price * quantity_kg
                        purchase_price_stats[source] += 1
                
                # Use daily-calculated revenue
                total_revenue = total_actual_revenue
                
                # Calculate warehouse and delivery costs (3 birr/kg and 2 birr/kg)
                total_warehouse_costs = 0.0
                total_delivery_costs = 0.0
                for item in daily_sales["items"]:
                    quantity_kg = float(item.get("quantity_sold", 0.0) or item.get("quantity_kg", 0.0))
                    if quantity_kg > 0:
                        total_warehouse_costs += quantity_kg * 3.0
                        total_delivery_costs += quantity_kg * 2.0
                
                # Calculate CM1, CM2, CM3
                cm1_amount = total_revenue - total_actual_cogs
                cm1_percent = (cm1_amount / total_revenue * 100) if total_revenue > 0 else 0
                
                cm2_amount = cm1_amount - total_warehouse_costs
                cm2_percent = (cm2_amount / total_revenue * 100) if total_revenue > 0 else 0
                
                cm3_amount = cm2_amount - total_delivery_costs
                cm3_percent = (cm3_amount / total_revenue * 100) if total_revenue > 0 else 0
                
                # Build result structure
                result = {
                    "margin_breakdown": {
                        "cm1_gross_profit": {
                            "amount": round(cm1_amount, 2),
                            "percentage": f"{cm1_percent:.2f}%",
                            "formula": "Revenue - Product COGS",
                            "description": "Revenue minus direct product costs (purchase price)."
                        },
                        "cm2_after_warehouse": {
                            "amount": round(cm2_amount, 2),
                            "percentage": f"{cm2_percent:.2f}%",
                            "formula": "CM1 - Warehouse Costs",
                            "description": "CM1 minus warehouse handling costs."
                        },
                        "cm3_net_profit": {
                            "amount": round(cm3_amount, 2),
                            "percentage": f"{cm3_percent:.2f}%",
                            "formula": "CM2 - Delivery Costs",
                            "description": "CM2 minus last-mile delivery costs, representing final profit."
                        },
                    },
                    "summary": {
                        "total_revenue": round(total_revenue, 2),
                        "product_cogs": round(total_actual_cogs, 2),
                        "warehouse_costs": round(total_warehouse_costs, 2),
                        "delivery_costs": round(total_delivery_costs, 2),
                        "final_profit": round(cm3_amount, 2),
                        "final_margin": f"{cm3_percent:.2f}%",
                    },
                    "period": date_range,
                    "purchase_price_metadata": {
                        "cogs_recalculated": True,
                        "calculation_method": "daily_transactions",
                        "purchase_price_sources": dict(purchase_price_stats),
                        "note": "Calculated on daily basis: COGS = purchase_price × quantity per day, then summed. Revenue from daily product orders."
                    }
                }
                
                return result
            else:
                # No daily data available
                return {
                    "margin_breakdown": {
                        "cm1_gross_profit": {"amount": 0, "percentage": "0%", "formula": "Revenue - Product COGS", "description": "Revenue minus direct product costs."},
                        "cm2_after_warehouse": {"amount": 0, "percentage": "0%", "formula": "CM1 - Warehouse Costs", "description": "CM1 minus warehouse handling costs."},
                        "cm3_net_profit": {"amount": 0, "percentage": "0%", "formula": "CM2 - Delivery Costs", "description": "CM2 minus last-mile delivery costs."},
                    },
                    "summary": {
                        "total_revenue": 0,
                        "product_cogs": 0,
                        "warehouse_costs": 0,
                        "delivery_costs": 0,
                        "final_profit": 0,
                        "final_margin": "0%",
                    },
                    "period": date_range,
                    "error": "No B2B data available for the specified date range"
                }
        except Exception as e:
            logger.warning(f"Could not calculate profit margins: {e}")
            return {
                "margin_breakdown": {
                    "cm1_gross_profit": {"amount": 0, "percentage": "0%", "formula": "Revenue - Product COGS", "description": "Revenue minus direct product costs."},
                    "cm2_after_warehouse": {"amount": 0, "percentage": "0%", "formula": "CM1 - Warehouse Costs", "description": "CM1 minus warehouse handling costs."},
                    "cm3_net_profit": {"amount": 0, "percentage": "0%", "formula": "CM2 - Delivery Costs", "description": "CM2 minus last-mile delivery costs."},
                },
                "summary": {
                    "total_revenue": 0,
                    "product_cogs": 0,
                    "warehouse_costs": 0,
                    "delivery_costs": 0,
                    "final_profit": 0,
                    "final_margin": "0%",
                },
                "period": date_range,
                "error": f"Failed to calculate profit margins: {str(e)}"
            }
    except Exception as e:
        logger.error(f"Error fetching B2B profit margin analysis: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch B2B profit margin analysis: {str(e)}")


@app.get("/api/b2b/cost-structure-analysis")
async def get_b2b_cost_structure_analysis(
    date_from: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="End date (YYYY-MM-DD)")
):
    """Get B2B cost structure analysis: COGS, warehouse, delivery costs. Calculated from daily product orders."""
    try:
        client = get_b2b_mcp_client()
        purchase_price_service = get_b2b_purchase_price_service()
        date_range = format_date_range(date_from, date_to)
        
        # Get daily product orders data
        daily_sales = await client.get_daily_sales_data(date_from, date_to)
        
        if not daily_sales or "items" not in daily_sales or len(daily_sales.get("items", [])) == 0:
            return {
                "cogs": 0,
                "warehouse_costs": 0,
                "delivery_costs": 0,
                "other_costs": 0,
                "total_costs": 0,
                "breakdown": [],
                "date_range": date_range,
                "error": "No B2B data available for the specified date range"
            }
        
        # Calculate costs from daily data
        total_cogs = 0.0
        total_warehouse = 0.0
        total_delivery = 0.0
        
        for item in daily_sales["items"]:
            product_name = item.get("product_name", "")
            order_date_str = item.get("order_date") or item.get("sale_date")
            quantity_kg = float(item.get("quantity_sold", 0.0) or item.get("quantity_kg", 0.0))
            
            if quantity_kg <= 0:
                continue
            
            # Parse order date for purchase price lookup
            try:
                if isinstance(order_date_str, str):
                    order_date = datetime.strptime(order_date_str.split("T")[0], "%Y-%m-%d").date()
                else:
                    order_date = order_date_str
            except:
                continue
            
            # Get purchase price for this order date (with next-day offset)
            purchase_price, _ = purchase_price_service.get_purchase_price_for_sale_date(
                product_name, order_date
            )
            
            # Calculate costs for this day
            if purchase_price > 0:
                total_cogs += purchase_price * quantity_kg
            total_warehouse += quantity_kg * 3.0
            total_delivery += quantity_kg * 2.0
        
        total_costs = total_cogs + total_warehouse + total_delivery
        
        return {
            "cogs": round(total_cogs, 2),
            "warehouse_costs": round(total_warehouse, 2),
            "delivery_costs": round(total_delivery, 2),
            "other_costs": 0,
            "total_costs": round(total_costs, 2),
            "breakdown": [
                {
                    "category": "Product COGS",
                    "amount": round(total_cogs, 2),
                    "percentage": round((total_cogs / total_costs * 100) if total_costs > 0 else 0, 2)
                },
                {
                    "category": "Warehouse Costs",
                    "amount": round(total_warehouse, 2),
                    "percentage": round((total_warehouse / total_costs * 100) if total_costs > 0 else 0, 2)
                },
                {
                    "category": "Delivery Costs",
                    "amount": round(total_delivery, 2),
                    "percentage": round((total_delivery / total_costs * 100) if total_costs > 0 else 0, 2)
                }
            ],
            "date_range": date_range
        }
    except Exception as e:
        logger.error(f"Error fetching B2B cost structure analysis: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch B2B cost structure analysis: {str(e)}")


@app.get("/api/b2b/product-profit-analysis")
async def get_b2b_product_profit_analysis(
    date_from: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="End date (YYYY-MM-DD)")
):
    """Get B2B per-product profit and volume analysis with actual purchase prices."""
    try:
        from datetime import datetime as dt
        from collections import defaultdict
        
        client = get_b2b_mcp_client()
        purchase_price_service = get_b2b_purchase_price_service()
        date_range = format_date_range(date_from, date_to)
        
        # First, try to get daily transaction-level data for accurate calculations
        daily_sales = await client.get_daily_sales_data(date_from, date_to)
        use_daily_data = daily_sales and "items" in daily_sales and len(daily_sales.get("items", [])) > 0
        
        if use_daily_data:
            # Calculate on daily basis: for each day, calculate revenue and COGS, then sum
            product_data: Dict[str, Dict[str, Any]] = defaultdict(lambda: {
                "product_name": "",
                "total_revenue": 0.0,
                "total_quantity_kg": 0.0,
                "order_ids": set(),
                "total_cogs": 0.0,
                "warehouse_costs": 0.0,
                "delivery_costs": 0.0,
                "purchase_price_sources": defaultdict(int),
                "days_with_data": set(),
                "order_count": 0
            })
            
            # Process each daily transaction
            for item in daily_sales["items"]:
                product_name = item.get("product_name") or item.get("product_id", "")
                if not product_name:
                    continue
                
                # Get daily transaction data from new MCP structure
                order_date_str = item.get("order_date") or item.get("sale_date") or item.get("date")
                quantity_kg = float(item.get("quantity_sold", 0.0) or item.get("quantity_kg", 0.0) or item.get("quantity", 0.0))
                # Use total_revenue from MCP (already calculated correctly per day)
                daily_revenue = float(item.get("total_revenue", 0.0) or item.get("revenue", 0.0))
                num_orders = int(item.get("number_of_orders", 0))
                
                if quantity_kg <= 0:
                    continue
                
                # Parse order date
                try:
                    if isinstance(order_date_str, str):
                        order_date = dt.strptime(order_date_str.split("T")[0], "%Y-%m-%d").date()
                    else:
                        order_date = order_date_str
                except:
                    continue
                
                # Get purchase price for this specific order date (with next-day offset)
                purchase_price, source = purchase_price_service.get_purchase_price_for_sale_date(
                    product_name, order_date
                )
                
                # Calculate daily COGS = purchase_price × quantity
                daily_cogs = purchase_price * quantity_kg if purchase_price > 0 else 0.0
                
                # Calculate daily warehouse and delivery costs (3 birr/kg and 2 birr/kg)
                daily_warehouse = quantity_kg * 3.0
                daily_delivery = quantity_kg * 2.0
                
                # Aggregate by product
                prod = product_data[product_name]
                prod["product_name"] = product_name
                prod["total_revenue"] += daily_revenue
                prod["total_quantity_kg"] += quantity_kg
                prod["total_cogs"] += daily_cogs
                prod["warehouse_costs"] += daily_warehouse
                prod["delivery_costs"] += daily_delivery
                prod["purchase_price_sources"][source] += 1
                prod["days_with_data"].add(order_date)
                # Track order count from number_of_orders field
                prod["order_count"] = prod.get("order_count", 0) + num_orders
        
        if not use_daily_data:
            # No daily data available - return empty (new MCP should always provide daily data)
            logger.warning("Daily product orders data not available from new MCP endpoint")
            return {
                "products": [],
                "summary": {
                    "total_products": 0,
                    "total_revenue": 0,
                    "total_profit": 0,
                    "avg_profit_margin": 0
                },
                "period": date_range,
                "error": "Daily product orders data not available from MCP endpoint"
            }
        
        # Convert to list and calculate final metrics
        products_list = []
        total_revenue = 0.0
        total_profit = 0.0
        
        for product_name, data in product_data.items():
            revenue = data["total_revenue"]
            cogs = data["total_cogs"]
            warehouse = data["warehouse_costs"]
            delivery = data["delivery_costs"]
            profit = revenue - cogs - warehouse - delivery
            margin = (profit / revenue * 100) if revenue > 0 else 0.0
            profit_per_kg = (profit / data["total_quantity_kg"]) if data["total_quantity_kg"] > 0 else 0.0
            
            # Calculate average purchase price used (weighted by quantity if daily data)
            # For aggregated data, this is the single price used
            # For daily data, this is the weighted average
            avg_purchase_price = (cogs / data["total_quantity_kg"]) if data["total_quantity_kg"] > 0 else 0.0
            
            # Calculate average selling price per kg (for display only - revenue is already correctly calculated)
            avg_selling_price = (revenue / data["total_quantity_kg"]) if data["total_quantity_kg"] > 0 else 0.0
            
            products_list.append({
                "product_name": product_name,
                "total_revenue": round(revenue, 2),
                "total_quantity_kg": round(data["total_quantity_kg"], 2),
                "total_orders": data.get("order_count", len(data["order_ids"])),
                "avg_selling_price": round(avg_selling_price, 2),
                "purchase_price_used": round(data.get("purchase_price_used", avg_purchase_price), 2),
                "total_cogs": round(cogs, 2),
                "warehouse_costs": round(warehouse, 2),
                "delivery_costs": round(delivery, 2),
                "total_profit": round(profit, 2),
                "profit_margin_percent": round(margin, 2),
                "profit_per_kg": round(profit_per_kg, 2),
                "days_with_data": len(data["days_with_data"]) if data["days_with_data"] else 1  # Default to 1 for aggregated data
            })
            
            total_revenue += revenue
            total_profit += profit
        
        # Sort by profit (descending)
        products_list.sort(key=lambda x: x["total_profit"], reverse=True)
        
        # Calculate summary
        avg_margin = (total_profit / total_revenue * 100) if total_revenue > 0 else 0.0
        
        return {
            "products": products_list,
            "summary": {
                "total_products": len(products_list),
                "total_revenue": round(total_revenue, 2),
                "total_profit": round(total_profit, 2),
                "avg_profit_margin": round(avg_margin, 2)
            },
            "period": date_range,
            "calculation_method": "daily_transactions",
            "note": "Calculated using daily transactions: revenue and COGS calculated per day (revenue = total_revenue per day, COGS = purchase_price × quantity per day), then summed across all days."
        }
        
    except Exception as e:
        logger.error(f"Error fetching B2B product profit analysis: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch B2B product profit analysis: {str(e)}")


@app.get("/api/b2b/revenue-trends")
async def get_b2b_revenue_trends(
    date_from: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    granularity: Optional[str] = Query("day", description="Granularity: day, week, month")
):
    """Get B2B revenue trends over time. Calculated from daily product orders."""
    try:
        from collections import defaultdict
        
        client = get_b2b_mcp_client()
        date_range = format_date_range(date_from, date_to)
        
        # Get daily product orders data
        daily_sales = await client.get_daily_sales_data(date_from, date_to)
        
        if not daily_sales or "items" not in daily_sales or len(daily_sales.get("items", [])) == 0:
            return {
                "trends": [],
                "period": date_range,
                "error": "No B2B data available for the specified date range"
            }
        
        # Aggregate revenue by date
        revenue_by_date = defaultdict(float)
        for item in daily_sales["items"]:
            order_date_str = item.get("order_date") or item.get("sale_date")
            revenue = float(item.get("total_revenue", 0.0) or item.get("revenue", 0.0))
            if order_date_str and revenue > 0:
                revenue_by_date[order_date_str] += revenue
        
        # Convert to list sorted by date
        trends = [
            {"date": date_str, "revenue": round(revenue, 2)}
            for date_str, revenue in sorted(revenue_by_date.items())
        ]
        
        return {
            "trends": trends,
            "period": date_range,
            "granularity": granularity
        }
    except Exception as e:
        logger.error(f"Error fetching B2B revenue trends: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch B2B revenue trends: {str(e)}")


@app.get("/api/b2b/cash-flow-analysis")
async def get_b2b_cash_flow_analysis(
    date_from: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="End date (YYYY-MM-DD)")
):
    """Get B2B cash flow analysis: cash in/out by payment method.
    
    Note: The new Product Orders MCP does not provide cash flow data.
    This endpoint returns a placeholder response indicating the feature is not available.
    """
    try:
        date_range = format_date_range(date_from, date_to)
        # The new Product Orders MCP only provides getDailyProductOrders
        # Cash flow analysis is not available from this MCP endpoint
        return {
            "cash_in": 0,
            "cash_out": 0,
            "net_cash_flow": 0,
            "payment_methods": {},
            "period": date_range,
            "error": "Cash flow analysis is not available from the Product Orders MCP. This feature requires payment method data which is not provided by the current MCP endpoint.",
            "note": "The Product Orders MCP only provides daily product order statistics (getDailyProductOrders). Cash flow data would require a different data source."
        }
    except Exception as e:
        logger.error(f"Error fetching B2B cash flow analysis: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch B2B cash flow analysis: {str(e)}")


@app.get("/api/b2b/customer-profitability")
async def get_b2b_customer_profitability(
    date_from: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="End date (YYYY-MM-DD)")
):
    """Get B2B customer profitability analysis: CLTV & margins per customer."""
    try:
        client = get_b2b_mcp_client()
        date_range = format_date_range(date_from, date_to)
        result = await client.call_tool("get_customer_profitability_analysis", {"date_range": date_range})
        return result
    except Exception as e:
        logger.error(f"Error fetching B2B customer profitability: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch B2B customer profitability: {str(e)}")


@app.get("/api/b2b/credit-risk-dashboard")
async def get_b2b_credit_risk_dashboard(
    date_from: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="End date (YYYY-MM-DD)")
):
    """Get B2B credit risk dashboard: credit stages & overdue."""
    try:
        client = get_b2b_mcp_client()
        date_range = format_date_range(date_from, date_to)
        result = await client.call_tool("get_credit_risk_dashboard", {"date_range": date_range})
        return result
    except Exception as e:
        logger.error(f"Error fetching B2B credit risk dashboard: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch B2B credit risk dashboard: {str(e)}")


@app.get("/api/b2b/payment-behavior")
async def get_b2b_payment_behavior(
    date_from: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="End date (YYYY-MM-DD)")
):
    """Get B2B payment behavior: payment history & DSO. Not available with current MCP endpoint."""
    return {
        "error": "Payment behavior analysis not available with the current Product Orders MCP endpoint",
        "note": "This feature requires additional customer payment data that is not provided by the current MCP.",
        "period": format_date_range(date_from, date_to)
    }

# Serve static files from frontend build directory (for Docker deployment)
# This must be added after all API routes
if FRONTEND_DIST.exists():
    # Serve static assets (JS, CSS, images, etc.)
    if (FRONTEND_DIST / "assets").exists():
        app.mount("/assets", StaticFiles(directory=str(FRONTEND_DIST / "assets")), name="assets")
    
    # Serve other static files (favicon, etc.) and handle SPA routing
    @app.get("/{path:path}")
    async def serve_frontend(path: str):
        """Serve frontend files, with fallback to index.html for SPA routing."""
        # Skip API routes (should already be handled by FastAPI, but just in case)
        if path.startswith("api/"):
            raise HTTPException(status_code=404, detail="API endpoint not found")
        
        # Try to serve the requested file
        file_path = FRONTEND_DIST / path
        if file_path.exists() and file_path.is_file():
            return FileResponse(file_path)
        
        # For SPA routing, return index.html for all non-API routes
        if (FRONTEND_DIST / "index.html").exists():
            return FileResponse(FRONTEND_DIST / "index.html")
        
        raise HTTPException(status_code=404, detail="Not found")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)
